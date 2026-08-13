# -*- coding: utf-8 -*-
"""Aba "Análise Base SEFAZ".

Le a SEFAZ.xlsx -- que NAO e a base do resto do painel -- e devolve as linhas ja
no formato que o gerar_painel.py consome (cabecalho + tuplas), para passarem
pelo mesmo `montar_aba`.

DUAS ABAS, UMA TELA
-------------------
A planilha tem duas abas com colunas completamente diferentes: `NFes SEFAZ`
(nota de produto, 386 colunas) e `NFS` (nota de servico, 139). Elas viram UMA
aba do painel, com a coluna **Origem** dizendo de onde cada linha veio -- e os
campos equivalentes empilhados na mesma coluna:

    emitente  <-> prestador          destinatario <-> tomador
    Vlr Total NF <-> Valor dos Servicos     Info Comp <-> Descricao do Servico

SO AS COLUNAS MARCADAS ENTRAM
-----------------------------
Das 386 colunas o usuario pintou de VERDE as que interessam (as outras estao em
azul). Sao 16 na aba de NF-e e 10 na de NFS-e -- e so essas viram coluna aqui.
A cor nao e lida do arquivo: ela mora nas listas COLUNAS_NFE/COLUNAS_NFS abaixo,
por POSICAO, do jeito que a SF1 tambem e lida. Motivo: nomes se repetem e vem
com grafia irregular ("Razão  Social Prestador" tem dois espacos), enquanto a
posicao e estavel. O NOME esperado viaja junto e e CONFERIDO na leitura: se a
planilha mudar de layout, a leitura para com o nome da coluna que saiu do lugar,
em vez de trazer dado errado calado.

⚠ A ABA DE NF-e E POR ITEM x DUPLICATA, nao por nota: 726 linhas para 155
chaves (uma nota chegou a ter 51 linhas). Isso e de proposito -- `Vlr Total
Item`, `CFOP`, `Venc Duplicata` e `Valor Duplicata` sao desses niveis, e o
usuario marcou as quatro. Por isso a identidade da linha soma chave + item +
duplicata: so com a chave, as 51 linhas de uma nota seriam uma marcacao so; so
com a chave e o item, uma nota de 1 item paga em 3x ainda colidiria (aconteceu
com 8 notas). A aba de NFS-e nao tem nada disso: 462 chaves para 462 linhas.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import re
import shutil
import tempfile
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

BASE_PADRAO = Path(
    r"C:\Users\lucas\OneDrive - Grupo S&D\Arquivos de Gabriella Karla Oliveira Milas - FINANCEIRO COMPARTILHADO"
    r"\LUCAS ABNER ARAUJO\AUTOMAÇÕES LUCAS\ANALISES BOLETOS\SEFAZ.xlsx"
)

# Prefixo do uuid, no mesmo espirito do "ITAU:" e do "BOL:": as marcacoes de
# todas as abas moram na mesma tabela, e chave repetida seria marcacao trocada.
PREFIXO = "SEFAZ:"

ABA_NFE = "NFes SEFAZ"
ABA_NFS = "NFS"

# Nomes de saida (o painel casa a coluna por este nome).
CHAVE = "Chave"
ORIGEM = "Origem"
CHAVE_NFE = "Chave NF-e"
NUM_NF = "Nº NF"
EMISSAO = "Emissão"
SAIDA = "Saída/Entrada"
TIPO_OP = "Tipo Operação"
# ⚠ NAO e o "Tipo Operação" acima: aquele diz entrada/saida, este e a natureza
# escrita na nota ("COMPRA PARA COMERCIALIZACAO", "PRESTACAO DE SERVICO"...). As
# duas abas trazem, com NOMES DIFERENTES -- "Nat. Operação" na NF-e (F) e
# "Natureza Operação" na NFS-e (P) --, e por isso a trava de nome e por aba.
NAT_OP = "Nat. Operação"
CFOP = "CFOP"
VLR_ITEM = "Vlr Item"
VLR_TOTAL = "Vlr Total"
VENC_DUP = "Venc Duplicata"
VLR_DUP = "Vlr Duplicata"
INFO = "Informações"
CNPJ_EMIT = "CNPJ Emitente"
NOME_EMIT = "Emitente"
FANTASIA = "Nome Fantasia"
CNPJ_DEST = "CNPJ Destinatário"
NOME_DEST = "Destinatário"

ALERTA = "Alerta"

CABECALHO = [CHAVE, ORIGEM, ALERTA, NUM_NF, EMISSAO, SAIDA, TIPO_OP, NAT_OP, CFOP,
             NOME_EMIT, CNPJ_EMIT, FANTASIA, NOME_DEST, CNPJ_DEST,
             VLR_ITEM, VLR_TOTAL, VENC_DUP, VLR_DUP, INFO, CHAVE_NFE]

# ------------------------------------------------------------------- alertas
# 13/08/2026, pedido do usuario. Tres avisos, todos no nivel da NOTA -- ele pediu
# "informando NF que entram na condicao", e nao a duplicata:
#   1) vence em 5 dias ou menos E nao esta lancada na SF1;
#   2) emitida com vencimento a menos de 15 dias da emissao;
#   3) emitida ha mais de 3 dias E nao esta lancada na SF1.
#
# ⚠ Quem sabe o que esta lancado e o cruzamento com a SF1, entao o conjunto
# `nao_lancadas` chega de FORA -- so as chaves. Assim este modulo nao precisa do
# vocabulario do cruzamento (e nao ha import circular: e ele que importa este).
# ⚠ NFS-e NAO TEM VENCIMENTO nesta base -- a aba `NFS` nao traz a coluna --,
# entao nela so o alerta 3 pode disparar. Limite do arquivo, nao do codigo.
# ⚠ Vencida tambem entra no alerta 1: "5 dias ou menos do vencimento" e uma
# barra de urgencia, e quem passou dela e mais urgente, nao menos.
# ⚠ O calculo e da GERACAO: o alerta envelhece ate a proxima rodada do painel.
DIAS_VENCE_PERTO = 5
DIAS_PRAZO_CURTO = 15
DIAS_SEM_LANCAR = 3

# Os prefixos sao contrato: o gerar_painel classifica o selo por eles (ALERTAS) e
# o resumo conta por eles. Mudar o texto aqui pede mudar os dois.
PREFIXO_VENCE = "VENCE EM ATÉ"
PREFIXO_SEM_LANCAR = "SEM LANÇAMENTO NA SF1"
PREFIXO_PRAZO = "PRAZO CURTO"


def _dias(n: int) -> str:
    return "1 dia" if abs(n) == 1 else f"{n} dias"


def vencimentos_por_nota(linhas: list[dict]) -> dict[str, list]:
    """{chave NF-e: vencimentos das duplicatas dela}. Nota sem duplicata fica fora."""
    mapa: dict[str, list] = {}
    for linha in linhas:
        if linha.get(VENC_DUP):
            mapa.setdefault(linha[CHAVE_NFE], []).append(linha[VENC_DUP])
    return mapa


def alertas_da_nota(emissao, vencimentos: list, nao_lancada: bool,
                    hoje: dt.date) -> list[str]:
    """Os avisos desta nota, do mais urgente para o menos.

    O vencimento considerado e o PRIMEIRO da nota: e ele que chega antes, tanto
    para a urgencia (alerta 1) quanto para o prazo concedido (alerta 2).
    """
    avisos = []
    venc = min((v for v in vencimentos if v), default=None)

    if venc and nao_lancada:
        faltam = (venc - hoje).days
        if faltam <= DIAS_VENCE_PERTO:
            quando = (f"venceu há {_dias(-faltam)}" if faltam < 0
                      else "vence hoje" if faltam == 0
                      else f"vence em {_dias(faltam)}")
            avisos.append(f"{PREFIXO_VENCE} {DIAS_VENCE_PERTO} DIAS E NÃO ESTÁ LANÇADA: "
                          f"{quando} ({venc.strftime('%d/%m/%Y')}) e não foi encontrada na SF1")

    if emissao and nao_lancada:
        parada = (hoje - emissao).days
        if parada > DIAS_SEM_LANCAR:
            avisos.append(f"{PREFIXO_SEM_LANCAR} HÁ MAIS DE {DIAS_SEM_LANCAR} DIAS: "
                          f"emitida em {emissao.strftime('%d/%m/%Y')}, há {_dias(parada)}")

    if emissao and venc:
        prazo = (venc - emissao).days
        if prazo < DIAS_PRAZO_CURTO:
            avisos.append(f"{PREFIXO_PRAZO}: emitida com vencimento {_dias(prazo)} depois da "
                          f"emissão ({emissao.strftime('%d/%m/%Y')} → {venc.strftime('%d/%m/%Y')})")
    return avisos


def alertas_por_nota(linhas: list[dict], nao_lancadas: set | None,
                     hoje: dt.date) -> dict[str, list[str]]:
    """{chave NF-e: avisos}. Uma vez por NOTA, ainda que ela tenha 51 linhas."""
    nao_lancadas = nao_lancadas or set()
    vencimentos = vencimentos_por_nota(linhas)
    por_nota: dict[str, list[str]] = {}
    for linha in linhas:
        chave = linha[CHAVE_NFE]
        if chave in por_nota:
            continue
        por_nota[chave] = alertas_da_nota(
            linha.get(EMISSAO), vencimentos.get(chave, []), chave in nao_lancadas, hoje)
    return por_nota


def contar_alertas(por_nota: dict[str, list[str]]) -> dict:
    """Quantas NOTAS em cada condicao -- e o que o aviso da aba mostra."""
    tem = lambda prefixo: sum(1 for avisos in por_nota.values()
                              if any(a.startswith(prefixo) for a in avisos))
    return {
        "vence_sem_lancar": tem(PREFIXO_VENCE),
        "sem_lancar": tem(PREFIXO_SEM_LANCAR),
        "prazo_curto": tem(PREFIXO_PRAZO),
        "com_alerta": sum(1 for avisos in por_nota.values() if avisos),
    }

# Selos da coluna Origem. Curtos: a celula corta o que nao couber.
NFE = "NF-e"
NFSE = "NFS-e"

# (letra na planilha, nome esperado no cabecalho, nome no painel).
# A letra e o que vale; o nome e a TRAVA -- ver o docstring.
COLUNAS_NFE = [
    ("C",  "Chave NF-e",         CHAVE_NFE),
    ("I",  "Num NF",             NUM_NF),
    ("J",  "Data Emissão",       EMISSAO),
    ("K",  "Data Saída/Entrada", SAIDA),
    ("L",  "Tipo Operação",      TIPO_OP),
    ("F",  "Nat. Operação",      NAT_OP),
    ("W",  "CFOP",               CFOP),
    ("AA", "Vlr Total Item",     VLR_ITEM),
    ("LM", "Venc Duplicata",     VENC_DUP),
    ("LN", "Valor Duplicata",    VLR_DUP),
    ("NJ", "Vlr Total NF",       VLR_TOTAL),
    ("LP", "Info Comp",          INFO),
    ("NK", "CNPJ Emitente",      CNPJ_EMIT),
    ("NM", "Nome Emitente",      NOME_EMIT),
    ("NN", "Nome Fantasia",      FANTASIA),
    ("NR", "CNPJ Dest",          CNPJ_DEST),
    ("NT", "Nome Dest",          NOME_DEST),
]

COLUNAS_NFS = [
    ("A",  "Chave NFS-e",             CHAVE_NFE),
    ("I",  "Número NFSe",             NUM_NF),
    ("H",  "Data de Emissão NFSe",    EMISSAO),
    ("P",  "Natureza Operação",       NAT_OP),
    ("Y",  "Descrição do Serviço",    INFO),
    ("BJ", "Valor dos Serviços",      VLR_TOTAL),
    ("EB", "CNPJ Prestador",          CNPJ_EMIT),
    ("EC", "Razão  Social Prestador", NOME_EMIT),
    ("EE", "Nome Fantasia Prestador", FANTASIA),
    ("DV", "CNPJ Tomador",            CNPJ_DEST),
    ("DW", "Razão Social Tomador",    NOME_DEST),
]

# Numero do item, so para a identidade da linha -- nao vira coluna na tela
# (o usuario nao pintou essa) mas e o que separa as 51 linhas de uma nota.
ITEM_NFE = ("P", "Item Nº")

# O texto integral vai para o quadro flutuante, nao para a celula. O teto existe
# porque `Info Comp` chega a 5.000 caracteres e a carteira inteira viaja para o
# navegador depois do login: sem corte, esta aba sozinha dobraria o download.
# O que passa disso vira "...", e a planilha continua tendo o texto completo.
TETO_TEXTO = 1200


def _sem_acento(texto: str) -> str:
    normalizado = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def _rotulo(valor) -> str:
    """Nome de coluna reduzido ao essencial, para conferir o cabecalho."""
    return re.sub(r"\s+", " ", _sem_acento(valor).replace("\xa0", " ")).strip().lower()


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (dt.datetime, dt.date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return re.sub(r"[ \t]+", " ", str(valor).replace("\xa0", " ")).strip()


def _data(valor):
    """Devolve `date` -- e o que faz o painel formatar dd/mm/aaaa e ordenar."""
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    bruto = _texto(valor)
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(bruto, formato).date()
        except ValueError:
            pass
    return None


def _numero(valor):
    """Numero cru; o painel formata em R$. Texto com virgula tambem entra."""
    if valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    limpo = re.sub(r"[^0-9,.\-]", "", str(valor))
    if "," in limpo:
        limpo = limpo.replace(".", "").replace(",", ".")
    try:
        return float(limpo)
    except ValueError:
        return None


def _uma_linha(texto: str) -> str:
    """Texto de varias linhas em uma so, com o teto aplicado."""
    junto = re.sub(r"\s*\n\s*", " · ", texto).strip(" ·")
    junto = re.sub(r"\s{2,}", " ", junto)
    return junto if len(junto) <= TETO_TEXTO else junto[:TETO_TEXTO].rstrip() + "…"


def nf_chave(valor) -> str:
    """NF so em digitos e sem zeros a esquerda -- e assim que ela casa com a SF1."""
    return re.sub(r"\D", "", _texto(valor)).lstrip("0")


def _abrir(caminho: Path):
    """Abre a planilha; se o Excel/OneDrive estiver segurando, usa uma copia."""
    try:
        return load_workbook(caminho, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        pasta = tempfile.mkdtemp(prefix="sefaz_")
        shutil.copy2(caminho, Path(pasta) / caminho.name)
        return load_workbook(Path(pasta) / caminho.name, read_only=True,
                             data_only=True), pasta


def _conferir(cabecalho: list, colunas: list, aba: str) -> None:
    """Cada coluna esperada continua onde estava? Se nao, para aqui.

    Trazer a coluna errada calado seria pior do que nao trazer nada: ninguem
    descobriria olhando a tela que o CNPJ do emitente virou outra coisa.
    """
    erros = []
    for letra, esperado, _destino in colunas:
        i = column_index_from_string(letra) - 1
        achado = _rotulo(cabecalho[i]) if i < len(cabecalho) else "(nao existe)"
        if achado != _rotulo(esperado):
            erros.append(f"{letra} deveria ser {esperado!r} e esta {achado!r}")
    if erros:
        raise RuntimeError(
            f"A aba {aba!r} da SEFAZ.xlsx mudou de layout:\n  - "
            + "\n  - ".join(erros)
            + "\nCorrija as posicoes em sefaz.py (COLUNAS_NFE/COLUNAS_NFS)."
        )


def _ler_aba(ws, colunas: list, origem: str, com_item: bool) -> list[dict]:
    brutas = ws.iter_rows(values_only=True)
    cabecalho = list(next(brutas, ()) or ())
    _conferir(cabecalho, colunas, ws.title)
    posicoes = [(column_index_from_string(l) - 1, destino) for l, _n, destino in colunas]
    item_i = column_index_from_string(ITEM_NFE[0]) - 1 if com_item else None

    linhas = []
    for bruta in brutas:
        if all(v in (None, "") for v in bruta):
            continue
        linha = {ORIGEM: origem}
        for i, destino in posicoes:
            valor = bruta[i] if i < len(bruta) else None
            if destino in (EMISSAO, SAIDA, VENC_DUP):
                linha[destino] = _data(valor)
            elif destino in (VLR_ITEM, VLR_TOTAL, VLR_DUP):
                linha[destino] = _numero(valor)
            elif destino == INFO:
                linha[destino] = _uma_linha(_texto(valor))
            else:
                linha[destino] = _texto(valor)
        # sem chave nao ha nota -- e sobra de formatacao no fim da planilha
        if not linha.get(CHAVE_NFE):
            continue
        item = _texto(bruta[item_i]) if item_i is not None and item_i < len(bruta) else ""
        # ⚠ A duplicata entra na identidade junto com o item: a aba de NF-e nao e
        # so por item, e por ITEM x DUPLICATA. Uma nota de 1 item paga em 3x vem
        # em 3 linhas identicas em tudo, menos no vencimento e no valor da
        # parcela -- so com o item, essas 3 seriam a MESMA marcacao (foi o que
        # aconteceu com 8 notas na primeira leitura).
        parcela = f"{linha.get(VENC_DUP) or ''}|{linha.get(VLR_DUP) or ''}"
        linha[CHAVE] = PREFIXO + hashlib.sha1(
            f"{origem}|{linha[CHAVE_NFE]}|{item}|{parcela}".encode("utf-8")).hexdigest()[:16]
        linhas.append(linha)
    return linhas


def ler(base: Path) -> list[dict]:
    """As duas abas da SEFAZ.xlsx, ja unificadas e com a Origem marcada."""
    wb, temporaria = _abrir(base)
    try:
        linhas = []
        if ABA_NFE in wb.sheetnames:
            linhas += _ler_aba(wb[ABA_NFE], COLUNAS_NFE, NFE, com_item=True)
        if ABA_NFS in wb.sheetnames:
            linhas += _ler_aba(wb[ABA_NFS], COLUNAS_NFS, NFSE, com_item=False)
        return linhas
    finally:
        wb.close()
        if temporaria:
            shutil.rmtree(temporaria, ignore_errors=True)


def montar_linhas(linhas: list[dict]) -> list[tuple]:
    """Linhas -> tuplas na ordem de CABECALHO, prontas para o `montar_aba`."""
    # Emissao mais recente primeiro: e a nota que acabou de chegar da SEFAZ que
    # ainda nao foi olhada. Sem data vai para o fim.
    ordenadas = sorted(linhas, key=lambda l: (l.get(EMISSAO) is None,
                                              -(l[EMISSAO].toordinal() if l.get(EMISSAO) else 0),
                                              l.get(NOME_EMIT, "")))
    return [tuple(l.get(coluna, "") for coluna in CABECALHO) for l in ordenadas]


def carregar(base: Path, nao_lancadas: set | None = None,
             hoje: dt.date | None = None) -> tuple[list[str], list[tuple], dict]:
    """Ponto de entrada: (cabecalho, linhas, resumo) para o gerador.

    `nao_lancadas` sao as chaves NF-e que o cruzamento com a SF1 nao achou. Sem
    elas os alertas 1 e 3 nao tem como existir (ninguem procurou), e so o de
    prazo curto -- que depende apenas da propria nota -- continua valendo.
    """
    linhas = ler(base)
    hoje = hoje or dt.date.today()
    # O alerta e da NOTA: a frase se repete nas linhas dela, como o Vlr Total ja
    # se repete. Sem isso, a linha da duplicata 2 nao avisaria nada.
    por_nota = alertas_por_nota(linhas, nao_lancadas, hoje)
    for linha in linhas:
        linha[ALERTA] = " · ".join(por_nota.get(linha[CHAVE_NFE], []))

    nfe = [l for l in linhas if l[ORIGEM] == NFE]
    nfs = [l for l in linhas if l[ORIGEM] == NFSE]
    resumo = {
        **contar_alertas(por_nota),
        "total": len(linhas),
        "nfe": len(nfe),
        "nfs": len(nfs),
        "notas_nfe": len({l[CHAVE_NFE] for l in nfe}),
        "notas_nfs": len({l[CHAVE_NFE] for l in nfs}),
        "com_duplicata": sum(1 for l in linhas if l.get(VLR_DUP)),
    }
    return list(CABECALHO), montar_linhas(linhas), resumo
