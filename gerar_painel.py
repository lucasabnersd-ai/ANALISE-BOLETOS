# -*- coding: utf-8 -*-
"""Gera o painel HTML da aba 'Associacoes Encontradas'.

Le TRATAMENTO PYTHON BOLETOS.xlsx e escreve um index.html unico, sem
dependencias externas: a mesma planilha, com as mesmas colunas e na mesma
ordem. O painel so acrescenta o que a planilha nao consegue mostrar -- o
destaque quando o boleto diverge do titulo e o botao de copiar UUID e linha
digitavel.

Uso:
    python gerar_painel.py [--base CAMINHO] [--saida CAMINHO]
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import shutil
import sys
import tempfile
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

BASE_PADRAO = Path(
    r"C:\Users\lucas\OneDrive - Grupo S&D\Arquivos de Gabriella Karla Oliveira Milas - FINANCEIRO COMPARTILHADO"
    r"\LUCAS ABNER ARAUJO\TRATAMENTO PYTHON BOLETOS.xlsx"
)
PASTA = Path(__file__).resolve().parent
SAIDA_PADRAO = PASTA / "PUBLICAR" / "index.html"
MODELO = PASTA / "painel_modelo.html"
# Fora do repositorio (.gitignore): a carteira para subir e o painel de uso local.
DADOS_JSON = PASTA / "DADOS" / "analise_boletos.json"
DEV = PASTA / "dev.html"

ABA_ASSOCIACOES = "Associações Encontradas"
ABA_CORRESPONDENCIAS = "Tratar Correspondências"
ABA_RESUMO = "Resumo"
ABA_PARCELAS = "NFs Múltiplas Parcelas"

# So a linha digitavel entra; o codigo de barras fica de fora (pedido do usuario).
IGNORAR = {"Código de Barras"}

# Coluna de marcacao: fica no painel mesmo vazia na planilha, virando caixa de
# selecao. O que ja estiver preenchido na base entra marcado.
COLUNA_CHECK = "CHECK (FEITO)"

MOEDA = {"Vlr.Titulo", "Valor Boleto", "Saldo", "Desconto", "Multa", "Juros",
         "Correcao", "Val Liq Baix",
         # nomes usados na aba Tratar Correspondencias
         "Valor Título (R$)", "Valor Boleto (R$)", "Diferença (R$)"}
DATA = {"Vencimento", "Vencto Real", "Vencimento Boleto", "DT Emissao",
        "DT Baixa", "DT Contab."}
INTEIRO = {"Score Match", "2º Score", "Margem", "Score"}

# Colunas com botao de copiar (o usuario cola no SE2 / no banco). O rotulo so
# aparece nas que sao SO_BOTAO; nas outras o proprio valor e o botao.
COPIAVEIS = {"Campo UUID": "UUID", "Linha Digitável": "LINHA",
             "Fornecedor": "", "No. Titulo": ""}

# Dessas o valor nem aparece: a coluna e so o botao de copiar (sao longas demais
# e o usuario nunca le, so cola).
SO_BOTAO = {"Campo UUID", "Linha Digitável"}

# Tipos de alerta que ganham selo proprio na coluna Alerta. A ordem importa:
# o texto de fatura tambem fala em "parcelas", entao FATURA e testado primeiro.
ALERTAS = ((("FATURA",), "FATURA"),
           (("MAIS DE UMA PARCELA", "OUTRAS PARCELAS", "PARCELA"), "PARCELA"))

# Confrontos que viram destaque na celula do boleto.
#   coluna destacada -> (coluna do titulo, tipo)
CONFRONTOS = {
    "Valor Boleto": ("Vlr.Titulo", "moeda"),
    "Vencimento Boleto": ("Vencimento", "data"),
}

# Visao Conferencia (unica visao do painel): subconjunto das colunas NA MESMA
# ORDEM DA BASE. Como a base ja poe titulo e boleto lado a lado, a unica coisa
# acrescentada e a coluna de diferenca (@delta_*), logo depois do par.
#   (rotulo, lado, grupo, largura relativa)
# Empresa pagadora e contraparte NAO entram aqui: elas sao colunas do drill
# (o usuario corrigiu isso em 04/08/2026). A tabela principal segue enxuta.
COMPACTA = [
    ("CHECK (FEITO)",     "",       "",           2.5),
    ("Campo UUID",        "",       "",           4),
    ("Filial",            "",       "",           3.5),
    ("Prefixo",           "",       "",           3.5),
    # A NF do boleto sai do lugar dela na base para ficar colada na do titulo
    # (pedido do usuario): e essa comparacao que ele faz o tempo todo.
    ("No. Titulo",        "titulo", "NF",         6.5),
    ("NF/Doc Boleto",     "boleto", "NF",         6),
    ("Parcela",           "titulo", "",           3),
    ("Fornecedor",        "titulo", "",           5),
    ("Razão Social",      "titulo", "",          13),
    ("Vlr.Titulo",        "titulo", "VALOR",      6.5),
    ("Valor Boleto",      "boleto", "VALOR",      6.5),
    ("@delta_valor",      "delta",  "VALOR",      7.5),
    ("Vencimento",        "titulo", "VENCIMENTO", 5.5),
    ("Vencto Real",       "titulo", "VENCIMENTO", 5.5),
    ("Vencimento Boleto", "boleto", "VENCIMENTO", 5.5),
    ("@delta_dias",       "delta",  "VENCIMENTO", 4.5),
    ("Status",            "",       "",           7),
    ("Alerta Fatura",     "",       "",           5),
    ("@tratativa",        "",       "",           4),
    ("Linha Digitável",   "",       "",           4),
]

# Aba "Tratar Correspondencias": MESMAS funcoes, colunas adaptadas ao que ela
# tem. Nao existe `Status` ali -- o que manda e o `Motivo Revisão`, e valor,
# diferenca e vencimento vem em colunas proprias.
COMPACTA_CORRESP = [
    ("CHECK (FEITO)",       "",       "",           2.5),
    ("Campo UUID",          "",       "",           4),
    ("Filial",              "",       "",           3.5),
    ("Prefixo",             "",       "",           3.5),
    ("No. Titulo",          "titulo", "NF",         6.5),
    ("NF/Doc Boleto",       "boleto", "NF",         6),
    ("Parcela",             "titulo", "",           3),
    ("Fornecedor",          "titulo", "",           5),
    ("Razão Social",        "titulo", "",          12),
    ("Fornecedor Boleto",   "boleto", "",          11),
    ("Valor Título (R$)",   "titulo", "VALOR",      6.5),
    ("Valor Boleto (R$)",   "boleto", "VALOR",      6.5),
    ("@delta_valor",        "delta",  "VALOR",      7.5),
    ("Vencimento",          "titulo", "VENCIMENTO", 5.5),
    ("Vencto Real",         "titulo", "VENCIMENTO", 5.5),
    ("Vencimento Boleto",   "boleto", "VENCIMENTO", 5.5),
    ("@delta_dias",         "delta",  "VENCIMENTO", 4.5),
    ("Motivo Revisão",      "",       "",           7),
    ("Alerta Fatura",       "",       "",           5),
    ("@tratativa",          "",       "",           4),
    ("Linha Digitável",     "",       "",           4),
]

# Confrontos e nomes de coluna mudam entre as abas; o resto do gerador e igual.
ABAS = {
    "associacoes": {
        "planilha": ABA_ASSOCIACOES,
        "nome": "Associações Encontradas",
        "cor": "azul",
        "compacta": COMPACTA,
        "confrontos": {"Valor Boleto": ("Vlr.Titulo", "moeda"),
                       "Vencimento Boleto": ("Vencimento", "data")},
        "col_status": "Status",
        "col_criterio": "Critério Match",
        "col_score": "Score Match",
    },
    "correspondencias": {
        "planilha": ABA_CORRESPONDENCIAS,
        "nome": "Tratar Correspondências",
        "cor": "roxo",
        "compacta": COMPACTA_CORRESP,
        "confrontos": {"Valor Boleto (R$)": ("Valor Título (R$)", "moeda"),
                       "Vencimento Boleto": ("Vencimento", "data")},
        "col_status": "Motivo Revisão",
        "col_criterio": "Critério",
        "col_score": "Score",
    },
}

# Cabecalho mais curto na Conferencia (o painel casa a coluna pela chave, nao
# pelo rotulo -- renomear aqui e seguro).
# Rotulos curtos: o cabecalho e uma linha so, entao nome comprido seria cortado.
ROTULOS_COMPACTA = {"@delta_valor": "Δ valor", "@delta_dias": "Δ dias",
                    "Alerta Fatura": "Alerta", "Campo UUID": "UUID",
                    "Linha Digitável": "Linha", "@tratativa": "Nota",
                    "CHECK (FEITO)": "OK", "No. Titulo": "Nº título",
                    "NF/Doc Boleto": "NF boleto", "Prefixo": "Pref.",
                    "Parcela": "Parc.", "Razão Social": "Razão social",
                    "Fornecedor": "Cód.", "Vlr.Titulo": "Vlr. título",
                    "Valor Boleto": "Vlr. boleto", "Vencimento": "Vencto",
                    "Vencto Real": "Vencto real", "Vencimento Boleto": "Venc. boleto",
                    # nomes da aba Tratar Correspondencias
                    "Valor Título (R$)": "Vlr. título", "Valor Boleto (R$)": "Vlr. boleto",
                    "Motivo Revisão": "Motivo", "Fornecedor Boleto": "Forn. boleto"}

LARGURAS = {
    "Campo UUID": 296, "Razão Social": 250, "Fornecedor Boleto": 200,
    "Nome Fornece": 175, "Critério Match": 280, "Alerta Fatura": 280,
    "Linha Digitável": 290, "Historico": 300, "CNPJ Boleto": 145,
    "No. Titulo": 104, "Status": 122, "Fonte Boleto": 135,
}
LARGURA_PADRAO = 108


def chave_de(rotulo: str) -> str:
    texto = unicodedata.normalize("NFKD", rotulo)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"_+", "_", re.sub(r"[^0-9a-z]+", "_", texto.lower())).strip("_")


def abrir_planilha(caminho: Path):
    """Abre a planilha; se estiver travada pelo Excel/OneDrive, usa uma copia."""
    try:
        return load_workbook(caminho, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        pasta = tempfile.mkdtemp(prefix="painel_boletos_")
        shutil.copy2(caminho, Path(pasta) / caminho.name)
        return load_workbook(Path(pasta) / caminho.name, read_only=True, data_only=True), pasta


def texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (dt.datetime, dt.date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    limpo = re.sub(r"\s+", " ", str(valor).replace("_x000D_", " ")).strip()
    return "" if limpo in ("/ /", "//") else limpo


def numero(valor):
    if valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    limpo = re.sub(r"[^0-9,.\-]", "", str(valor))
    if "," in limpo:
        limpo = limpo.replace(".", "").replace(",", ".")
    try:
        return float(limpo)
    except ValueError:
        return None


def data_iso(valor):
    if isinstance(valor, (dt.datetime, dt.date)):
        return valor.strftime("%Y-%m-%d")
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(texto(valor), formato).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def moeda_br(valor) -> str:
    if valor is None:
        return ""
    inteiro = f"{valor:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"R$ {inteiro}"


def formatar_linha_digitavel(valor) -> str:
    """Deixa a linha no formato lido pelo banco: 5.5 5.6 5.6 1 14."""
    digitos = re.sub(r"\D", "", texto(valor))
    if len(digitos) != 47:
        return texto(valor)
    return " ".join([
        f"{digitos[0:5]}.{digitos[5:10]}", f"{digitos[10:15]}.{digitos[15:21]}",
        f"{digitos[21:26]}.{digitos[26:32]}", digitos[32], digitos[33:47],
    ])


def tipo_da_coluna(rotulo: str) -> str:
    if rotulo in MOEDA:
        return "moeda"
    if rotulo in DATA:
        return "data"
    if rotulo in INTEIRO:
        return "inteiro"
    return "texto"


def montar_colunas(cabecalho: list[str], usadas: set[str]) -> list[dict]:
    return [
        {
            "chave": chave_de(rotulo),
            "rotulo": rotulo,
            "tipo": tipo_da_coluna(rotulo),
            "copiar": COPIAVEIS.get(rotulo, ""),
            "check": rotulo == COLUNA_CHECK,
            "largura": 58 if rotulo == COLUNA_CHECK else LARGURAS.get(rotulo, LARGURA_PADRAO),
        }
        for rotulo in cabecalho
        if rotulo and rotulo not in IGNORAR and (rotulo in usadas or rotulo == COLUNA_CHECK)
    ]


def delta_bruto(origem: dict, coluna_boleto: str, confrontos: dict):
    """Diferenca numerica boleto - titulo (reais ou dias); None se nao da para comparar."""
    coluna_titulo, tipo = confrontos[coluna_boleto]
    if tipo == "moeda":
        valor_t, valor_b = numero(origem.get(coluna_titulo)), numero(origem.get(coluna_boleto))
        return None if None in (valor_t, valor_b) else round(valor_b - valor_t, 2)

    data_t, data_b = data_iso(origem.get(coluna_titulo)), data_iso(origem.get(coluna_boleto))
    if not (data_t and data_b):
        return None
    return (dt.date.fromisoformat(data_b) - dt.date.fromisoformat(data_t)).days


def diferenca(origem: dict, coluna_boleto: str, confrontos: dict) -> str:
    """Texto curto da divergencia entre o boleto e o titulo ('' se conferem)."""
    delta = delta_bruto(origem, coluna_boleto, confrontos)
    if not delta:
        return ""
    if confrontos[coluna_boleto][1] == "moeda":
        return ("+" if delta > 0 else "−") + moeda_br(abs(delta))
    return f"{delta:+d} {'dia' if abs(delta) == 1 else 'dias'}"


def ler_aba(wb, cfg: dict):
    """Le uma aba da base no formato do painel. As duas abas passam por aqui --
    o que muda entre elas vem do cfg (colunas, confrontos, nomes de campo)."""
    ws = wb[cfg["planilha"]]
    linhas = list(ws.iter_rows(values_only=True))
    cabecalho = [texto(c) for c in linhas[0]]
    brutas = [b for b in linhas[1:] if not all(v in (None, "") for v in b)]
    compacta = montar_compacta(cabecalho, cfg["compacta"], cfg)
    confrontos = cfg["confrontos"]
    moeda_boleto = next(c for c, (_, t) in confrontos.items() if t == "moeda")

    # Uma coluna 100% vazia nas linhas associadas nao ajuda ninguem a conferir.
    usadas = {
        rotulo for i, rotulo in enumerate(cabecalho)
        if rotulo and any(texto(b[i]) for b in brutas)
    }
    colunas = montar_colunas(cabecalho, usadas)

    registros = []
    for bruta in brutas:
        origem = dict(zip(cabecalho, bruta))
        celulas, ordem = {}, {}

        for coluna in colunas:
            rotulo, chave, tipo = coluna["rotulo"], coluna["chave"], coluna["tipo"]
            valor = origem.get(rotulo)
            if tipo == "moeda":
                bruto = numero(valor)
                celulas[chave], ordem[chave] = moeda_br(bruto), bruto
            elif tipo == "data":
                celulas[chave], ordem[chave] = texto(valor), data_iso(valor) or ""
            elif tipo == "inteiro":
                bruto = numero(valor)
                celulas[chave] = "" if bruto is None else str(int(bruto))
                ordem[chave] = bruto
            elif rotulo == "Linha Digitável":
                celulas[chave] = formatar_linha_digitavel(valor)
                ordem[chave] = re.sub(r"\D", "", texto(valor))
            else:
                celulas[chave] = ordem[chave] = texto(valor)

        registros.append({
            "c": celulas,
            "o": ordem,
            "uuid": texto(origem.get("Campo UUID")),
            "feito": bool(texto(origem.get(COLUNA_CHECK))),
            # valor que o botao copia (o exibido tem pontuacao de leitura)
            "copia": {chave_de("Linha Digitável"): re.sub(r"\D", "", texto(origem.get("Linha Digitável")))},
            "difere": {chave_de(col): diferenca(origem, col, confrontos) for col in confrontos
                       if diferenca(origem, col, confrontos)},
            # deltas numericos: alimentam a visao Conferencia e a ordenacao
            "delta": {"valor": delta_bruto(origem, moeda_boleto, confrontos),
                      "dias": delta_bruto(origem, "Vencimento Boleto", confrontos)},
            # `forte` so existe nas Associacoes; em Tratar Correspondencias todo
            # titulo esta em revisao, entao o selo nunca e verde.
            "forte": "FORTE" in texto(origem.get(cfg["col_status"])).upper(),
            "alerta": montar_alerta(origem, cfg),
            "match": montar_match(origem, cfg),
            "busca": " ".join(texto(v) for v in origem.values() if texto(v)).lower(),
        })

    chave_score = chave_de(cfg["col_score"])
    registros.sort(key=lambda r: -(numero(r["o"].get(chave_score)) or 0))
    return colunas, compacta, registros


def montar_compacta(cabecalho: list[str], definicao: list, cfg: dict) -> list[dict]:
    """Colunas da visao Conferencia, respeitando a ordem em que estao na base."""
    ordem_base = {rotulo: i for i, rotulo in enumerate(cabecalho) if rotulo}
    colunas = []
    for rotulo, lado, grupo, largura in definicao:
        virtual = rotulo.startswith("@")
        # O check e recurso do painel, nao coluna da base: entra mesmo onde a
        # planilha nao tem "CHECK (FEITO)" -- e o caso de Tratar Correspondencias.
        if not virtual and rotulo != COLUNA_CHECK and rotulo not in ordem_base:
            continue  # a base mudou de nome: melhor faltar a coluna do que errar
        colunas.append({
            "chave": rotulo.lstrip("@") if virtual else chave_de(rotulo),
            "rotulo": ROTULOS_COMPACTA.get(rotulo, rotulo),
            "tipo": ("tratativa" if rotulo == "@tratativa" else "delta") if virtual
                    else tipo_da_coluna(rotulo),
            "lado": lado,
            "grupo": grupo,
            "largura": largura,
            # o painel decide o desenho pelo PAPEL, nao pelo nome da coluna --
            # em Tratar Correspondencias o "status" chama Motivo Revisão.
            "papel": ("status" if rotulo == cfg["col_status"]
                      else "alerta" if rotulo == "Alerta Fatura" else ""),
            "copiavel": rotulo in COPIAVEIS,
            "copiar": COPIAVEIS.get(rotulo, ""),
            "so_botao": rotulo in SO_BOTAO,
            "check": rotulo == COLUNA_CHECK,
        })
    return colunas


def tipo_alerta(valor) -> str:
    """PARCELA / FATURA / ALERTA -- rotulo curto do selo ('' quando nao ha alerta)."""
    texto_alerta = texto(valor).upper()
    if not texto_alerta:
        return ""
    for marcadores, rotulo in ALERTAS:
        if any(m in texto_alerta for m in marcadores):
            return rotulo
    return "ALERTA"


def montar_match(origem: dict, cfg: dict) -> dict:
    """Resumo do porque o boleto casou com o titulo -- vira o quadro que abre ao
    passar o mouse na celula de Status.

    O criterio da base e uma frase longa com as evidencias somadas por ' + ';
    aqui ela e quebrada em itens para ler de bate-pronto. A parte depois de
    'REVISAO:' (so em Tratar Correspondencias) vira o motivo em separado.
    """
    bruto = texto(origem.get(cfg["col_criterio"]))
    evidencias, revisao = bruto, ""
    if "REVISAO:" in bruto.upper():
        corte = re.split(r"\|?\s*REVIS[ÃA]O:\s*", bruto, maxsplit=1, flags=re.IGNORECASE)
        evidencias, revisao = corte[0].strip(" |"), (corte[1] if len(corte) > 1 else "").strip()

    itens = [p.strip() for p in re.split(r"\s*\+\s*", evidencias) if p.strip()]
    return {
        "status": texto(origem.get(cfg["col_status"])),
        "itens": itens,
        "revisao": revisao or texto(origem.get("Motivo Revisão")),
        "fonte": texto(origem.get("Fonte Boleto")),
        "score": texto(origem.get(cfg["col_score"])),
        "score2": texto(origem.get("2º Score")),
        "margem": texto(origem.get("Margem")),
    }


def montar_alerta(origem: dict, cfg: dict) -> dict:
    """Tipo do alerta + a NF que o drill de parcelas abre.

    A NF vem do TEXTO do alerta ("... - NF 000013303: ...") e nao da coluna
    NF/Doc Boleto: quem escreveu o alerta sabia de qual NF estava falando, e a
    coluna as vezes traz dois numeros no mesmo campo.
    """
    bruto = texto(origem.get("Alerta Fatura"))
    tipo = tipo_alerta(bruto)
    if not tipo:
        return {"tipo": "", "nf": ""}
    achada = re.search(r"\bNF\s+([0-9]+)", bruto, re.IGNORECASE)
    nf = nf_chave(achada.group(1)) if achada else nf_chave(origem.get("NF/Doc Boleto"))
    # O quadro que aparece ao passar o mouse: o que a base tem para justificar
    # o alerta, sem obrigar o usuario a abrir o Excel.
    return {
        "tipo": tipo,
        "nf": nf,
        "texto": bruto,
        "criterio": texto(origem.get(cfg["col_criterio"])),
        "fonte": texto(origem.get("Fonte Boleto")),
        "score": texto(origem.get(cfg["col_score"])),
        "score2": texto(origem.get("2º Score")),
        "margem": texto(origem.get("Margem")),
    }


def nf_chave(valor) -> str:
    """NF sem zeros a esquerda -- e o unico jeito de casar as duas abas."""
    return re.sub(r"\D", "", texto(valor)).lstrip("0")


def ler_parcelas(wb) -> dict[str, list[dict]]:
    """Boletos da aba 'NFs Multiplas Parcelas', agrupados por NF.

    A NF aparece em dois lugares e nem sempre no mesmo: em parte das linhas ela
    esta na coluna 'NF (9 Digitos)', em outras essa coluna vem VAZIA e a NF so
    existe dentro do 'Nosso Numero (Ref.)', no formato 0013303-01. Ignorar o
    segundo caminho deixaria a NF 13303 (3 boletos) fora do drill.
    """
    if ABA_PARCELAS not in wb.sheetnames:
        return {}

    linhas = list(wb[ABA_PARCELAS].iter_rows(values_only=True))
    if not linhas:
        return {}
    cabecalho = [texto(c) for c in linhas[0]]

    por_nf: dict[str, list[dict]] = {}
    for bruta in linhas[1:]:
        origem = dict(zip(cabecalho, bruta))
        nosso = texto(origem.get("Nosso Número (Ref.)"))
        chave = nf_chave(origem.get("NF (9 Dígitos)")) or nf_chave(nosso.split("-")[0])
        if not chave:
            continue
        digitos = re.sub(r"\D", "", texto(origem.get("Linha Digitável")))
        por_nf.setdefault(chave, []).append({
            # NF como a base mostra (9 digitos); quando a coluna vem vazia,
            # cai para a chave resolvida pelo Nosso Numero.
            "nf": texto(origem.get("NF (9 Dígitos)")) or chave,
            "parcela": texto(origem.get("Parcela")),
            "venc": texto(origem.get("Data Vencimento")),
            "emissao": texto(origem.get("Data Emissão")),
            "valor": moeda_br(numero(origem.get("Valor (R$)"))),
            "valor_n": numero(origem.get("Valor (R$)")) or 0.0,
            "fornecedor": texto(origem.get("Favorecido/Fornecedor")),
            "empresa": texto(origem.get("Empresa Pagadora")),
            "nosso": nosso,
            "linha": formatar_linha_digitavel(origem.get("Linha Digitável")),
            "copia": digitos,
        })

    for boletos in por_nf.values():
        boletos.sort(key=lambda b: (b["parcela"], b["venc"]))
    return por_nf


def ler_resumo(wb) -> dict:
    resumo = {}
    for chave, valor in wb[ABA_RESUMO].iter_rows(values_only=True):
        rotulo = texto(chave)
        if rotulo and valor is not None:
            resumo[rotulo] = valor if isinstance(valor, (int, float)) else texto(valor)
    return resumo


def gerar(base: Path, saida: Path) -> dict:
    if not base.exists():
        raise FileNotFoundError(f"Base nao encontrada: {base}")
    if not MODELO.exists():
        raise FileNotFoundError(f"Modelo do painel nao encontrado: {MODELO}")

    wb, temporaria = abrir_planilha(base)
    try:
        lidas = {}
        for ident, cfg in ABAS.items():
            if cfg["planilha"] not in wb.sheetnames:
                continue  # aba renomeada na base: melhor faltar do que quebrar
            lidas[ident] = ler_aba(wb, cfg)
        resumo = ler_resumo(wb)
        parcelas = ler_parcelas(wb)
    finally:
        wb.close()
        if temporaria:
            shutil.rmtree(temporaria, ignore_errors=True)

    def do_resumo(rotulo):
        valor = resumo.get(rotulo, 0)
        return int(valor) if isinstance(valor, (int, float)) else 0

    abas = []
    for ident, cfg in ABAS.items():
        if ident not in lidas:
            continue
        # `colunas` nao vai para o HTML: o painel so desenha `compacta`. A lista
        # continua servindo para montar as celulas de cada registro.
        _colunas, compacta, registros = lidas[ident]
        abas.append({
            "id": ident,
            "nome": cfg["nome"],
            "cor": cfg["cor"],
            "compacta": compacta,
            "linhas": registros,
            "com_alerta": sum(1 for r in registros if r["c"].get("alerta_fatura")),
            "divergentes": sum(1 for r in registros if r["difere"]),
            # So as NFs que algum alerta de parcela realmente abre -- nao adianta
            # levar todas as NFs da aba se o painel so tem 3 alertas.
            "parcelas": {nf: parcelas[nf] for nf in
                         {r["alerta"]["nf"] for r in registros
                          if r["alerta"]["tipo"] == "PARCELA" and r["alerta"]["nf"]}
                         if nf in parcelas},
        })

    dados = {
        "gerado_em": resumo.get("Gerado em") or dt.date.today().strftime("%d/%m/%Y"),
        "atualizado_em": dt.datetime.now().strftime("%d/%m/%Y às %H:%M"),
        "sem_codigo": do_resumo("Títulos sem código de barras"),
        "abas": abas,
    }

    carga = json.dumps(dados, ensure_ascii=False, separators=(",", ":"))
    modelo = MODELO.read_text(encoding="utf-8")

    # Publicacao: o HTML sai SEM a carteira (mesmo modelo dos outros paineis --
    # o Pages serve publicamente mesmo com o repo privado). A carteira vai para
    # um JSON a parte, que o supa.js entrega so depois do login.
    saida.parent.mkdir(parents=True, exist_ok=True)
    # O supa.js entra SO no publicado: no dev.html a carteira ja vem embutida e
    # uma tela de login local so atrapalharia. O ?v= e cache-bust -- sem ele o
    # navegador serve o supa.js velho depois de qualquer ajuste.
    versao = dt.datetime.now().strftime("%Y%m%d%H%M")
    saida.write_text(
        modelo.replace("<!--__SUPA__-->", f'<script src="supa.js?v={versao}"></script>'),
        encoding="utf-8")
    conferir_sem_dados(saida, dados)

    DADOS_JSON.parent.mkdir(parents=True, exist_ok=True)
    DADOS_JSON.write_text(carga, encoding="utf-8")

    # Uso local: dev.html com tudo embutido, fora do repositorio (.gitignore).
    DEV.write_text(modelo.replace("/*__DADOS__*/null", carga), encoding="utf-8")
    return dados


# Pedacos que so podem existir no dev.html. Se aparecerem no publicado, a
# geracao para -- e a mesma trava que o painel do Geraldo tem.
def conferir_sem_dados(arquivo: Path, dados: dict) -> None:
    html = arquivo.read_text(encoding="utf-8")
    suspeitos = []
    for aba in dados["abas"]:
        for registro in aba["linhas"]:
            for valor in (registro["uuid"], registro["copia"].get(chave_de("Linha Digitável"), "")):
                if valor and valor in html:
                    suspeitos.append(valor)
    if "/*__DADOS__*/null" not in html:
        suspeitos.append("marcador /*__DADOS__*/null ausente")
    if suspeitos:
        arquivo.unlink(missing_ok=True)
        raise RuntimeError(
            f"{arquivo.name} sairia com dados dentro ({len(suspeitos)} ocorrencia(s), "
            f"ex.: {suspeitos[0][:40]}). Arquivo apagado -- nada foi publicado."
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base", type=Path, default=BASE_PADRAO)
    parser.add_argument("--saida", type=Path, default=SAIDA_PADRAO)
    args = parser.parse_args()

    try:
        dados = gerar(args.base, args.saida)
    except Exception as exc:  # noqa: BLE001 - mensagem amigavel no .cmd
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1

    for aba in dados["abas"]:
        print(f"{aba['nome']}: {len(aba['linhas'])} linhas | {len(aba['compacta'])} colunas"
              f" | com alerta: {aba['com_alerta']} | boleto difere: {aba['divergentes']}"
              f" | NFs com drill: {len(aba['parcelas'])}")
    print(f"Publicavel (sem dados): {args.saida}")
    print(f"Carteira para subir:    {DADOS_JSON}")
    print(f"Painel local (com dados): {DEV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
