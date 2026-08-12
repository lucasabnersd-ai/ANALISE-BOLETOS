# -*- coding: utf-8 -*-
"""Aba "NFs pendentes Itau - conferencia contas a pagar".

Le a planilha COPIAR E COLAR BOLETOS PENDENTES.xlsx -- que NAO e a base do
resto do painel -- e devolve as linhas ja no formato que o gerar_painel.py
consome (cabecalho + tuplas), para passarem pelo mesmo `montar_aba`.

DUAS COISAS QUE ESTE MODULO RESOLVE E A PLANILHA NAO
-----------------------------------------------------
1. A planilha e um COPIAR E COLAR de varias telas do DDA do Itau, uma por
   empresa/carteira. Por isso ela vem com a linha de CABECALHO REPETIDA no meio
   dos dados, com blocos separados por linhas em branco e, quando a pessoa cola
   duas vezes, com o mesmo boleto duas vezes. Nada disso pode virar linha do
   painel.

2. Ela e um RETRATO DO DIA: amanha a pessoa cola a tela nova por cima e o que
   foi pago simplesmente some. Ler so a planilha faria o painel esquecer o
   boleto no dia seguinte. Por isso existe o HISTORICO (arquivo JSON, fora do
   repositorio): cada boleto entra UMA VEZ, guarda a data em que apareceu pela
   primeira vez e continua no painel mesmo depois de sair da planilha -- ai com
   a situacao "SAIU DA BASE", que na pratica quer dizer "provavelmente pago".
   E o que o usuario pediu: "puxar sempre dessa base diariamente os novos
   boletos, sem repetir os antigos".

A CHAVE de cada boleto (o que decide "ja vi este") e
    CNPJ + numero do documento + vencimento + valor ate o vencimento
O valor usado e o "Ate venc.", nunca o "A pagar": o segundo cresce com juros e
multa depois do vencimento, e o mesmo boleto entraria de novo todo dia.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
import shutil
import tempfile
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

BASE_PADRAO = Path(
    r"C:\Users\lucas\OneDrive - Grupo S&D\Arquivos de Gabriella Karla Oliveira Milas - FINANCEIRO COMPARTILHADO"
    r"\LUCAS ABNER ARAUJO\COPIAR E COLAR BOLETOS PENDENTES.xlsx"
)

# Prefixo do uuid, no mesmo espirito do "BOL:" das NFs nao associadas: evita
# colidir com os UUIDs de verdade das outras abas na tabela de marcacoes.
PREFIXO = "ITAU:"

# Nomes de saida das colunas. Nao sao os da planilha por acaso:
#   "Venc." vira "Vencimento" porque o painel inteiro (ordem padrao, filtro de
#   periodo, formatacao de data) procura essa chave -- com "Venc." a aba abriria
#   fora de ordem e o filtro de vencimento nao a enxergaria.
VENCIMENTO = "Vencimento"
PAGADOR = "Pagador/Agregado"
BENEFICIARIO = "Beneficiário"
DOCUMENTO = "Nº Doc."
CNPJ = "CPF/CNPJ"
ATE_VENC = "Até Venc."
A_PAGAR = "A Pagar"
TIPO = "Tipo de Boleto"
# A observacao da planilha vira o ALERTA do painel: e sempre um aviso do banco
# ("este boleto sofreu alteracao/instrucao por parte do beneficiario"), o mesmo
# papel que a coluna de alerta cumpre nas outras abas.
ALERTA = "Alerta Fatura"
SITUACAO = "Situação"
ENTROU = "Entrou em"
VISTO = "Visto em"
CHAVE = "Chave"

CABECALHO = [CHAVE, SITUACAO, ENTROU, VISTO, PAGADOR, BENEFICIARIO, CNPJ,
             DOCUMENTO, VENCIMENTO, ATE_VENC, A_PAGAR, TIPO, ALERTA]

# Colunas da planilha de origem, por nome normalizado. A planilha e colada a
# mao: aceitar variacoes de acento/maiuscula evita quebrar por um detalhe de
# digitacao. A ordem desta lista tambem e o layout padrao do DDA, usado como
# ultimo recurso quando o cabecalho nao e reconhecido.
ORIGEM = [
    (PAGADOR,      ("pagador/agregado", "pagador", "pagador agregado")),
    (BENEFICIARIO, ("beneficiario", "beneficiario/cedente", "cedente")),
    (CNPJ,         ("cpf/cnpj", "cnpj/cpf", "cnpj", "cpf")),
    (VENCIMENTO,   ("venc.", "venc", "vencimento", "data de vencimento")),
    (DOCUMENTO,    ("no doc.", "n doc.", "no doc", "n doc", "numero do documento",
                    "no documento", "doc.", "doc")),
    (ATE_VENC,     ("ate venc.", "ate venc", "ate o vencimento", "valor ate venc.")),
    (A_PAGAR,      ("a pagar", "valor a pagar", "total a pagar")),
    (TIPO,         ("tipo de boleto", "tipo boleto", "tipo")),
    (ALERTA,       ("observacoes", "observacao", "obs", "obs.")),
]

# Textos curtos de proposito: eles viram SELO na tabela, e a celula corta o que
# nao couber -- "PENDENTE NO DDA" aparecia como "PENDENT...".
SITUACAO_PENDENTE = "PENDENTE"
SITUACAO_SAIU = "SAIU DA BASE"

# O Itau devolve o pagador truncado em 30 caracteres e escrito de um jeito
# diferente a cada tela: "S & D FLORESTAL LOGISTICA LTDA", "S E D FLORESTAL
# LOGISTICA LTDA", "SED FLORESTAL LOGISTICA LTDA", "S - D FLORESTAL...". Sao a
# MESMA empresa, e sem juntar elas o filtro por empresa vira 16 botoes para 4
# empresas de verdade. A palavra do meio e o que identifica -- o prefixo "S&D" e
# justamente a parte que o banco escreve de todo jeito.
EMPRESAS = (("BIOENERGIA",   "S&D FLORESTAL BIOENERGIA"),
            ("BIOFLORESTAS", "S&D FLORESTAL BIOFLORESTAS"),
            ("LOGISTICA",    "S&D FLORESTAL LOGISTICA"),
            ("BRASMAC",      "BRASMAC MAQUINAS E VEICULOS"))


def empresa_pagadora(valor) -> str:
    """Nome canonico da empresa pagadora; o texto original quando nao da para
    identificar (o banco as vezes trunca ate sobrar so "S")."""
    bruto = _texto(valor)
    procurar = _sem_acento(bruto).upper()
    for marcador, nome in EMPRESAS:
        if marcador in procurar:
            return nome
    return bruto


def _sem_acento(texto: str) -> str:
    normalizado = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def _rotulo(valor) -> str:
    """Nome de coluna reduzido ao essencial, para casar cabecalho."""
    return re.sub(r"\s+", " ", _sem_acento(valor).replace("\xa0", " ")).strip().lower()


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (dt.datetime, dt.date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return re.sub(r"\s+", " ", str(valor).replace("\xa0", " ")).strip()


def _data(valor):
    """Devolve `date` -- e o que faz o painel formatar dd/mm/aaaa e ordenar."""
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    bruto = _texto(valor)
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(bruto, formato).date()
        except ValueError:
            pass
    return None


def _digitos(valor) -> str:
    return re.sub(r"\D", "", _texto(valor))


def chave_do_boleto(linha: dict) -> str:
    """Identidade do boleto. Mesma chave = mesmo boleto, ja visto antes.

    O valor que entra e o "Ate venc.", nao o "A pagar": depois do vencimento o
    segundo cresce sozinho (juros/multa) e o boleto voltaria como novo todo dia.
    """
    venc = _data(linha.get(VENCIMENTO))
    partes = [
        _digitos(linha.get(CNPJ)),
        _rotulo(linha.get(DOCUMENTO)),
        venc.isoformat() if venc else "",
        _digitos(linha.get(ATE_VENC)),
    ]
    return PREFIXO + hashlib.sha1("|".join(partes).encode("utf-8")).hexdigest()[:16]


def _mapear_colunas(cabecalho: list) -> dict:
    """Posicao de cada coluna da planilha. Cai para o layout padrao do DDA
    quando o cabecalho nao e reconhecido (planilha colada sem a primeira linha)."""
    rotulos = {_rotulo(c): i for i, c in enumerate(cabecalho) if _texto(c)}
    posicoes = {}
    for indice, (destino, aliases) in enumerate(ORIGEM):
        achado = next((rotulos[a] for a in aliases if a in rotulos), None)
        posicoes[destino] = achado if achado is not None else indice
    return posicoes


def _linha_de_cabecalho(bruta: tuple, posicoes: dict) -> bool:
    """A planilha traz o cabecalho repetido no meio (um por bloco colado)."""
    achados = 0
    for destino, aliases in ORIGEM:
        indice = posicoes[destino]
        if indice < len(bruta) and _rotulo(bruta[indice]) in aliases:
            achados += 1
    return achados >= 2


def ler_planilha(base: Path) -> list[dict]:
    """Linhas uteis da planilha, ja sem cabecalho repetido, vazio e duplicata.

    Duplicata dentro do MESMO arquivo (a pessoa colou o bloco duas vezes) segue
    a regra do resto do painel: vale a PRIMEIRA aparicao.
    """
    wb, temporaria = _abrir(base)
    try:
        ws = wb[wb.sheetnames[0]]
        brutas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
        if temporaria:
            shutil.rmtree(temporaria, ignore_errors=True)

    if not brutas:
        return []

    posicoes = _mapear_colunas(list(brutas[0]))
    linhas, vistas = [], set()
    for bruta in brutas:
        if all(v in (None, "") for v in bruta):
            continue
        if _linha_de_cabecalho(bruta, posicoes):
            continue

        linha = {}
        for destino in (d for d, _ in ORIGEM):
            indice = posicoes[destino]
            valor = bruta[indice] if indice < len(bruta) else None
            linha[destino] = _data(valor) if destino == VENCIMENTO else _texto(valor)

        # sem beneficiario e sem valor nao ha boleto -- e sobra de formatacao
        if not linha[BENEFICIARIO] and not _digitos(linha[ATE_VENC]):
            continue

        chave = chave_do_boleto(linha)
        if chave in vistas:
            continue
        vistas.add(chave)
        linha[CHAVE] = chave
        linhas.append(linha)
    return linhas


def _abrir(caminho: Path):
    """Abre a planilha; se o Excel/OneDrive estiver segurando, usa uma copia."""
    try:
        return load_workbook(caminho, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        pasta = tempfile.mkdtemp(prefix="pendentes_itau_")
        shutil.copy2(caminho, Path(pasta) / caminho.name)
        return load_workbook(Path(pasta) / caminho.name, read_only=True, data_only=True), pasta


def _ler_historico(arquivo: Path) -> dict:
    if not arquivo.exists():
        return {"versao": 1, "boletos": {}}
    try:
        guardado = json.loads(arquivo.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        # Historico ilegivel nao pode virar historico VAZIO em silencio: seria
        # o painel esquecendo tudo e reapresentando cada boleto como novo.
        raise RuntimeError(
            f"{arquivo.name} esta ilegivel. Nada foi gerado -- restaure o arquivo "
            "(ou apague-o de proposito, sabendo que todo boleto voltara como novo)."
        )
    guardado.setdefault("boletos", {})
    return guardado


def atualizar_historico(base: Path, arquivo: Path, hoje: dt.date | None = None) -> tuple[dict, dict]:
    """Cruza a planilha de hoje com o historico. NADA e removido do historico.

    Devolve (historico, resumo).
    """
    hoje = hoje or dt.date.today()
    dia = hoje.isoformat()

    historico = _ler_historico(arquivo)
    boletos = historico["boletos"]
    da_planilha = ler_planilha(base)

    novos, revistos, voltaram = 0, 0, 0
    presentes = set()

    for linha in da_planilha:
        chave = linha[CHAVE]
        presentes.add(chave)
        venc = linha[VENCIMENTO]
        gravavel = dict(linha)
        gravavel[VENCIMENTO] = venc.isoformat() if venc else ""

        antigo = boletos.get(chave)
        if antigo is None:
            boletos[chave] = {"entrou_em": dia, "visto_em": dia, "saiu_em": None,
                              "linha": gravavel}
            novos += 1
            continue

        # Ja conhecido: a linha e atualizada (o "A pagar" anda com juros), mas a
        # data de entrada e de quando ele apareceu pela primeira vez -- e o que
        # responde "ha quanto tempo este boleto esta parado".
        if antigo.get("saiu_em"):
            voltaram += 1
        antigo["linha"] = gravavel
        antigo["visto_em"] = dia
        antigo["saiu_em"] = None
        revistos += 1

    # Sumiu da planilha: sai da lista de pendentes, mas fica no painel com a
    # data em que saiu. Quem ja estava fora nao tem a data reescrita.
    sairam = 0
    for chave, guardado in boletos.items():
        if chave in presentes or guardado.get("saiu_em"):
            continue
        guardado["saiu_em"] = dia
        sairam += 1

    historico["versao"] = 1
    historico["atualizado_em"] = dt.datetime.now().isoformat(timespec="seconds")
    historico["base"] = str(base)
    resumo = {"na_planilha": len(da_planilha), "novos": novos, "revistos": revistos,
              "voltaram": voltaram, "sairam_hoje": sairam, "total": len(boletos),
              "pendentes": sum(1 for b in boletos.values() if not b.get("saiu_em"))}
    return historico, resumo


def gravar_historico(historico: dict, arquivo: Path) -> None:
    """Grava em arquivo temporario e so entao substitui: uma queda no meio da
    escrita nao pode deixar o historico pela metade."""
    arquivo.parent.mkdir(parents=True, exist_ok=True)
    provisorio = arquivo.with_suffix(arquivo.suffix + ".novo")
    provisorio.write_text(json.dumps(historico, ensure_ascii=False, indent=1),
                          encoding="utf-8")
    provisorio.replace(arquivo)


def montar_linhas(historico: dict) -> list[tuple]:
    """Historico -> tuplas na ordem de CABECALHO, prontas para o `montar_aba`."""
    registros = []
    for chave, guardado in historico["boletos"].items():
        linha = guardado.get("linha") or {}
        saiu = guardado.get("saiu_em")
        venc = _data(linha.get(VENCIMENTO))
        registros.append({
            CHAVE: chave,
            SITUACAO: SITUACAO_SAIU if saiu else SITUACAO_PENDENTE,
            ENTROU: _data(guardado.get("entrou_em")),
            VISTO: _data(saiu or guardado.get("visto_em")),
            # o historico guarda o texto do banco; a tela mostra o nome limpo
            PAGADOR: empresa_pagadora(linha.get(PAGADOR, "")),
            BENEFICIARIO: linha.get(BENEFICIARIO, ""),
            CNPJ: linha.get(CNPJ, ""),
            DOCUMENTO: linha.get(DOCUMENTO, ""),
            VENCIMENTO: venc,
            ATE_VENC: linha.get(ATE_VENC, ""),
            A_PAGAR: linha.get(A_PAGAR, ""),
            TIPO: linha.get(TIPO, ""),
            ALERTA: linha.get(ALERTA, ""),
        })

    # pendentes primeiro, e dentro de cada grupo o vencimento mais antigo antes
    registros.sort(key=lambda r: (r[SITUACAO] == SITUACAO_SAIU,
                                  r[VENCIMENTO] or dt.date.max))
    return [tuple(r[coluna] for coluna in CABECALHO) for r in registros]


def carregar(base: Path, arquivo_historico: Path,
             hoje: dt.date | None = None) -> tuple[list[str], list[tuple], dict]:
    """Ponto de entrada: le a planilha do dia, atualiza o historico e devolve
    (cabecalho, linhas, resumo) para o gerador."""
    historico, resumo = atualizar_historico(base, arquivo_historico, hoje)
    gravar_historico(historico, arquivo_historico)
    return list(CABECALHO), montar_linhas(historico), resumo
