# -*- coding: utf-8 -*-
"""Confere na SE2 quais titulos do painel ainda precisam de boleto.

Pedido do usuario (13/08/2026): a cada atualizacao do painel, os titulos das
abas de titulo do TOTVS sao conferidos contra a SE2. So continua na fila o
titulo EM ABERTO e SEM boleto lancado. Duas situacoes tiram o titulo dali:

    BAIXA   DT Baixa preenchida -- o titulo ja foi pago, nao ha o que lancar.
    BOLETO  Cod.Barras ou Linha Dig. preenchidos -- o boleto ja esta dentro do
            titulo, que era exatamente o trabalho do painel. E o caso do
            "rodou, alguem lancou, some na rodada seguinte".

⚠ A conferencia e SEMPRE contra a SE2 de verdade, NUNCA contra as colunas
"Codigo de Barras"/"Linha Digitavel" da base do painel: aquelas descrevem o
boleto SUGERIDO (vem do DDA/Itau), nao o que esta gravado no titulo. Confundir
as duas faria o painel esconder justamente o que ainda esta por fazer.

Aqui so se produz o veredito, uuid a uuid. Quem esconde e o navegador
(`filtrarSE2` no painel_modelo.html), porque e la que as marcacoes existem: o
titulo ja tratado por alguem CONTINUA aparecendo, na sub-aba Tratados, com o
selo do que aconteceu. Nada e apagado do banco.

Desde 13/08/2026 este mesmo passeio pela SE2 tambem traz o TIPO PGTO do titulo
(coluna IK), que o painel mostra ao lado do tipo de pagamento real informado a
mao -- de graca, porque a planilha ja esta aberta aqui.

Uso como modulo:
    conferencia = verificacao_se2.ler()          # None se a base nao existir
    conferencia.olhar(uuid)                      # None se o uuid nao esta la
    conferencia.tipos                            # vocabulario do Tipo Pgto
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

BASE_PADRAO = Path(
    r"C:\Users\lucas\OneDrive - Grupo S&D\Arquivos de Gabriella Karla Oliveira Milas - FINANCEIRO COMPARTILHADO"
    r"\LUCAS ABNER ARAUJO\BASES GENERICOS\SE2 - POSIÇÃO DIARIA.xlsx"
)

# A SE2 tem 23 mil linhas x 259 colunas: openpyxl leva ~30s para varrer isso, e
# a base so muda quando alguem exporta de novo. O cache guarda o veredito pronto
# e vale enquanto o arquivo tiver a MESMA data de modificacao E o mesmo tamanho
# -- exportacao nova muda os dois. Fica em DADOS\, que o .gitignore ja barra.
CACHE = Path(__file__).resolve().parent / "DADOS" / "verificacao_se2_cache.json"

# ⚠ Versao do FORMATO do cache, dentro da assinatura. Em 13/08/2026 o veredito
# ganhou o "t" (Tipo Pgto). Sem este numero, o cache gravado pela versao
# anterior -- mesma planilha, mesma data, mesmo tamanho -- seria aceito como
# valido e a coluna nova nasceria vazia, sem erro nenhum aparecer. Toda vez que
# `situacoes` mudar de forma, este numero sobe.
VERSAO_CACHE = 2

ABA = "SE2"
COL_UUID = "Campo UUID"
COL_BAIXA = "DT Baixa"
COL_BARRAS = "Cod.Barras"
COL_DIGITAVEL = "Linha Dig."
# Tipo de pagamento do titulo -- coluna IK da SE2 (a mesma que o painel de fluxo
# usa). Procurada pelo NOME, nunca pela posicao: o TOTVS ja inseriu coluna no
# meio dessa base antes. Ao contrario das quatro acima, esta e OPCIONAL: nao
# achar so deixa a coluna do painel vazia, nao derruba a conferencia.
COL_TIPO_PGTO = "Tipo Pgto"

# Codigo de barras tem 44 digitos e linha digitavel 47 (48 nos boletos de
# concessionaria). Conferido na base inteira: ou o campo vem vazio, ou vem
# completo -- nao ha meio termo. O piso esta aqui so para um "0" solto ou um
# resto de digitacao nao valer como boleto lancado.
MIN_DIGITOS = 20

# Data vazia no TOTVS nao e celula vazia: vem a mascara "  /  /    ".
VAZIO_DATA = re.compile(r"^[\s/:.-]*$")

# A SE2 exportada tem 259 colunas e nomes repetidos ("Status" e "Usuario"
# aparecem duas vezes). As quatro que interessam sao unicas -- por isso da para
# procurar por NOME aqui, ao contrario da SF1, que tem de ser lida por posicao.
def _indices(cabecalho: list[str]) -> dict[str, int]:
    faltando = []
    indices = {}
    for chave, rotulo in (("uuid", COL_UUID), ("baixa", COL_BAIXA),
                          ("barras", COL_BARRAS), ("digitavel", COL_DIGITAVEL)):
        if rotulo in cabecalho:
            indices[chave] = cabecalho.index(rotulo)
        else:
            faltando.append(rotulo)
    if faltando:
        raise KeyError("colunas ausentes na SE2: " + ", ".join(faltando))
    # Opcional: sem ela a conferencia continua inteira, so o tipo fica vazio.
    if COL_TIPO_PGTO in cabecalho:
        indices["tipo"] = cabecalho.index(COL_TIPO_PGTO)
    return indices


def _digitos(valor) -> str:
    return re.sub(r"\D", "", "" if valor is None else str(valor))


def _tipo_pgto(bruta, col: dict[str, int]) -> str:
    """Tipo Pgto do titulo em caixa alta ("" quando a coluna nao existe/vem vazia)."""
    i = col.get("tipo")
    if i is None or i >= len(bruta) or bruta[i] is None:
        return ""
    return re.sub(r"\s+", " ", str(bruta[i]).strip()).upper()


def _baixado(valor) -> str:
    """Data da baixa em dd/mm/aaaa, ou "" se o titulo esta em aberto."""
    if valor is None:
        return ""
    if hasattr(valor, "strftime"):
        return valor.strftime("%d/%m/%Y")
    texto = str(valor).strip()
    # "  /  /    " e a mascara de data vazia; qualquer coisa com digito e baixa
    return "" if VAZIO_DATA.match(texto) or not any(c.isdigit() for c in texto) else texto


class Conferencia:
    """O que a SE2 diz sobre cada titulo, pronto para consulta por uuid."""

    def __init__(self, arquivo: Path, situacoes: dict[str, dict],
                 tipos: list[str] | None = None):
        self.arquivo = arquivo
        self.situacoes = situacoes
        # Vocabulario do campo Tipo Pgto, do mais usado para o menos: e a lista
        # que vira o menu do "tipo real" no painel. Sai da propria SE2 de
        # proposito -- so vale escolher o que o TOTVS de fato usa.
        self.tipos = tipos or []
        carimbo = dt.datetime.fromtimestamp(arquivo.stat().st_mtime)
        self.salva_em = carimbo
        self.lido_em = carimbo.strftime("%d/%m/%Y às %H:%M")
        self.dias = (dt.datetime.now() - carimbo).days

    @property
    def total(self) -> int:
        return len(self.situacoes)

    def olhar(self, uuid: str) -> dict | None:
        """`{"m", "d", "dig", "t"}` do titulo, ou None quando ele nao esta na SE2.

        None NAO e "esta tudo certo": e "nao sei". Quem chama tem de deixar o
        titulo no painel nesse caso -- esconder por falta de informacao seria
        sumir com trabalho de verdade (titulo de outra empresa, base exportada
        pela metade, uuid em branco no TOTVS).
        """
        if not uuid:
            return None
        return self.situacoes.get(str(uuid).strip().upper())


def _assinatura(arquivo: Path) -> list:
    dados = arquivo.stat()
    return [VERSAO_CACHE, str(arquivo), int(dados.st_mtime), dados.st_size]


def _do_cache(arquivo: Path) -> dict | None:
    if not CACHE.exists():
        return None
    try:
        guardado = json.loads(CACHE.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return None            # cache corrompido: le a planilha de novo
    if guardado.get("assinatura") != _assinatura(arquivo):
        return None
    if not guardado.get("situacoes"):
        return None
    return guardado


def _guardar_cache(arquivo: Path, situacoes: dict, tipos: list[str]) -> None:
    try:
        CACHE.parent.mkdir(parents=True, exist_ok=True)
        CACHE.write_text(json.dumps({"assinatura": _assinatura(arquivo),
                                     "situacoes": situacoes, "tipos": tipos},
                                    ensure_ascii=False, separators=(",", ":")),
                         encoding="utf-8")
    except OSError:
        pass                   # sem cache o painel so fica mais lento


def ler(caminho: Path | None = None, usar_cache: bool = True) -> Conferencia | None:
    """Le a SE2 inteira. Devolve None quando a base nao esta na maquina.

    Sem a base o painel segue como sempre foi -- e melhor mostrar um titulo a
    mais do que esconder um que ainda precisa de boleto.
    """
    arquivo = Path(caminho) if caminho else BASE_PADRAO
    if not arquivo.exists():
        return None

    if usar_cache:
        guardado = _do_cache(arquivo)
        if guardado is not None:
            return Conferencia(arquivo, guardado["situacoes"],
                               guardado.get("tipos"))

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        ws = wb[ABA] if ABA in wb.sheetnames else wb[wb.sheetnames[0]]
        linhas = ws.iter_rows(values_only=True)
        cabecalho = ["" if c is None else str(c).strip() for c in next(linhas)]
        col = _indices(cabecalho)
        if "tipo" not in col:
            print(f"AVISO: a SE2 nao tem a coluna \"{COL_TIPO_PGTO}\" -- a coluna"
                  " de tipo de pagamento do painel fica vazia.", file=sys.stderr)

        situacoes: dict[str, dict] = {}
        quantos_tipos: Counter = Counter()
        for bruta in linhas:
            uuid = bruta[col["uuid"]] if col["uuid"] < len(bruta) else None
            if not uuid:
                continue
            baixa = _baixado(bruta[col["baixa"]])
            barras = _digitos(bruta[col["barras"]])
            digitavel = _digitos(bruta[col["digitavel"]])
            tem_boleto = max(len(barras), len(digitavel)) >= MIN_DIGITOS
            # Baixa vence boleto: titulo pago sai da fila de qualquer jeito, e
            # dizer "boleto lancado" de um titulo ja pago confunde quem le.
            if baixa:
                motivo, detalhe = "BAIXA", baixa
            elif tem_boleto:
                motivo, detalhe = "BOLETO", digitavel or barras
            else:
                motivo, detalhe = "", ""
            tipo = _tipo_pgto(bruta, col)
            if tipo:
                quantos_tipos[tipo] += 1
            situacoes[str(uuid).strip().upper()] = {
                "m": motivo, "d": detalhe,
                # digitos da linha digitavel gravada no titulo: e com isto que o
                # gerador descobre se lancaram o MESMO boleto que o painel
                # indicou ou um outro
                "dig": digitavel if motivo == "BOLETO" else "",
                # tipo de pagamento gravado no titulo (coluna IK)
                "t": tipo,
            }
    finally:
        wb.close()

    # do mais usado para o menos: e assim que a lista vira menu no painel
    tipos = [t for t, _ in quantos_tipos.most_common()]
    if usar_cache:
        _guardar_cache(arquivo, situacoes, tipos)
    return Conferencia(arquivo, situacoes, tipos)
