# -*- coding: utf-8 -*-
"""A manifestação do destinatário (MESMA PREMISSA.xlsx), colada nas notas da SEFAZ.

O que esta base acrescenta ao painel e uma coisa que ele nao tinha: o que a
empresa DISSE sobre a nota que originou o boleto -- deu ciencia, confirmou,
desconheceu, ou nao manifestou nada. Boleto a pagar de nota que a propria
empresa marcou como `DesconhecimentoOperacao` ou `NaoRealizada` e o caso que
esta base existe para pegar.

⚠ ESTE MODULO SO ENRIQUECE, NUNCA ACRESCENTA NOTA AO PAINEL. Medido em
01/09/2026: das 1.380 linhas do arquivo, 643 sao notas de TERCEIROS para a
BRACELL, em que a S&D so aparece como transportadora -- nota que o grupo nao
paga. Elas entram no arquivo porque o DF-e do TOTVS puxa tudo que cita o CNPJ do
grupo em qualquer papel. Como o cruzamento e um de-para por chave em cima das
notas que a SEFAZ ja trouxe, essas 643 se excluem sozinhas e nao precisa filtro
nenhum. Se um dia alguem usar este arquivo como FONTE de linhas, elas entram.

A chave e confiavel, e isso foi medido, nao suposto (01/09/2026):
  - 481 das 483 notas NF-e da SEFAZ estao aqui; as 2 de fora sao de 01/09,
    fora do recorte do arquivo (que vai de 01/08 a 31/08);
  - `ValorNF` bate com o `Vlr Total NF` da SEFAZ em 413 de 413 notas cruzadas,
    zero divergencia -- prova independente de que a chave casa a nota certa;
  - zero intersecao com NFS-e, e o correto: manifestacao nao existe para nota
    de servico.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

import sefaz

# ⚠ Derivado da pasta DESTE arquivo, e nao do `sefaz.BASE_PADRAO`: aquele tem um
# fallback para a maquina do Lucas quando a SEFAZ.xlsx nao esta la, e herdar o
# `.parent` dele faria esta base ser procurada num caminho de outro PC. Mesmo
# molde do sefaz.py: a pasta ANALISES BOLETOS, um nivel acima desta.
_AQUI = Path(__file__).resolve()
BASE_PADRAO = _AQUI.parents[1] / "MESMA PREMISSA.xlsx"

# ---------------------------------------------------------------- o arquivo
# ⚠ O ARQUIVO NAO E UMA PLANILHA: e um CSV embrulhado em .xlsx. Uma aba so, com
# UMA coluna, e o CSV inteiro dentro dela. O nome da aba e um GUID
# ('1ddb190e-f372-433a-9963-3797f6e') que muda a cada exportacao -- por isso a
# aba e sempre a PRIMEIRA, e nunca procurada por nome. Se um dia o export vier
# como planilha de verdade, o leitor aceita as duas formas (ver `_tabela`).
COLUNAS = ("NomeFilial", "Numero", "Serie", "DataEmissao", "Chave", "Emissor",
           "CnpjCpfEmitente", "Destinatario", "CnpjCpfDestinatario", "ValorNF",
           "Ambiente", "StatusManifestacao", "NaturezaOperacao",
           "JustificativaManifestacao")

CHAVE = "Chave"
STATUS = "StatusManifestacao"
FILIAL = "NomeFilial"
EMISSAO = "DataEmissao"
JUSTIFICATIVA = "JustificativaManifestacao"

# ⚠ O vocabulario e o da ORIGEM, sem traducao -- mesma regra do `Status da Nota`
# no sefaz.py. Inventar rotulo bonito aqui faria um status novo do TOTVS chegar
# na tela como se fosse conhecido, ou sumir num `else`.
SEM_MANIFESTACAO = "SemManifestacao"
CIENCIA = "CienciaOperacao"
CONFIRMACAO = "ConfirmacaoOperacao"
DESCONHECIMENTO = "DesconhecimentoOperacao"
NAO_REALIZADA = "NaoRealizada"

# Os dois que interessam ao painel de boletos: a empresa disse que nao reconhece
# a operacao, e mesmo assim existe titulo/boleto para pagar.
GRAVES = (DESCONHECIMENTO, NAO_REALIZADA)

# ⚠ NAO SAO STATUS: sao o que NOS dizemos quando a base nao fala da nota. Ficam
# graficamente diferentes dos valores da origem (travessao/caixa alta) de
# proposito -- quem olha a coluna tem de saber, sem perguntar, se aquilo veio do
# TOTVS ou e uma conclusao nossa.
NAO_SE_APLICA = "— não se aplica (NFS-e)"
FORA_DO_PERIODO = "— fora do período do arquivo"
NAO_ENCONTRADA = "NÃO ENCONTRADA"
SEM_BASE = "— base de manifestação não lida"


def _abrir(caminho: Path):
    """Abre o arquivo; se o Excel/OneDrive estiver segurando, usa uma copia.

    Mesmo remedio do sefaz.py, e pela mesma razao: este arquivo vive aberto na
    tela de quem acabou de exportar.
    """
    try:
        return load_workbook(caminho, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        pasta = tempfile.mkdtemp(prefix="manifesto_")
        shutil.copy2(caminho, Path(pasta) / caminho.name)
        return load_workbook(Path(pasta) / caminho.name, read_only=True,
                             data_only=True), pasta


def _tabela(ws) -> list[list[str]]:
    """As linhas do arquivo, venha ele como CSV numa coluna ou como planilha.

    ⚠ O CSV E LIDO COM `csv.reader`, NUNCA COM `split(',')`. Medido em
    01/09/2026: 21 das 1.380 linhas tem virgula DENTRO de um campo (entre
    aspas), e o split parte essas linhas em 15 ou 16 pedacos -- o que joga todos
    os campos seguintes uma casa para o lado. A chave da linha 'quebrada'
    passaria a ser o nome do emissor, e o cruzamento erraria calado.
    """
    if ws.max_column == 1:
        cruas = [c[0] for c in ws.iter_rows(values_only=True)]
        texto = "\n".join(str(c) for c in cruas if c is not None and str(c).strip())
        return [linha for linha in csv.reader(io.StringIO(texto)) if linha]
    tabela = []
    for bruta in ws.iter_rows(values_only=True):
        if all(v in (None, "") for v in bruta):
            continue
        tabela.append(["" if v is None else str(v).strip() for v in bruta])
    return tabela


def _conferir_cabecalho(cabecalho: list[str], base: Path) -> None:
    faltando = [c for c in COLUNAS if c not in cabecalho]
    if faltando:
        raise RuntimeError(
            f"{base.name} mudou de layout: faltam as colunas "
            + " / ".join(repr(c) for c in faltando) + "\n"
            "  o arquivo tem: " + " / ".join(repr(c) for c in cabecalho) + "\n"
            "Corrija COLUNAS no manifestacao.py."
        )


def _data(texto: str):
    """'04-08-2026T11:03:48AM-03:00' -> date(2026, 8, 4).

    So os 10 primeiros caracteres interessam: o resto e hora com AM/PM e fuso, e
    nada aqui e decidido por hora.
    """
    try:
        return dt.datetime.strptime(str(texto)[:10], "%d-%m-%Y").date()
    except (ValueError, TypeError):
        return None


def ler(base: Path | None = None) -> tuple[dict, dict]:
    """(de-para {chave NF-e: registro}, resumo). Uma entrada por NOTA.

    ⚠ A MESMA CHAVE VEM MAIS DE UMA VEZ. Medido em 01/09/2026: 56 chaves vem
    duas vezes, variando `StatusManifestacao` E `NomeFilial` -- a mesma nota
    vista por duas filiais, uma que ja manifestou e outra que nao. Nao ha
    duplicata pura. Se o join fosse feito sem resolver isso, essas 56 notas
    duplicariam as linhas da SEFAZ.
    A regra: qualquer status vence `SemManifestacao`, porque `SemManifestacao` e
    ausencia de resposta, e nao uma resposta. Dois status REAIS em conflito
    param a rodada -- nunca aconteceu (os 56 casos sao todos `X` + sem), e no
    dia em que acontecer e para alguem olhar, nao para o codigo escolher calado.
    """
    base = base or BASE_PADRAO
    if not base.exists():
        raise RuntimeError(
            f"A base de manifestação não foi encontrada:\n  {base}\n"
            "Sem ela a coluna de manifestação não tem como ser preenchida. "
            "Confira se o OneDrive sincronizou, ou corrija BASE_PADRAO no "
            "manifestacao.py."
        )
    wb, temporaria = _abrir(base)
    try:
        # A PRIMEIRA aba, sempre: o nome dela e um GUID que muda todo export.
        tabela = _tabela(wb[wb.sheetnames[0]])
    finally:
        wb.close()
        if temporaria:
            shutil.rmtree(temporaria, ignore_errors=True)

    if not tabela:
        raise RuntimeError(f"{base.name} está vazio — nada para cruzar.")
    cabecalho = [str(c).strip() for c in tabela[0]]
    _conferir_cabecalho(cabecalho, base)
    registros = [dict(zip(cabecalho, linha)) for linha in tabela[1:]]

    # ⚠ A chave tem 44 digitos e NAO cabe num float. Se alguem abrir este arquivo
    # e salvar como xlsx de verdade, o Excel transforma a coluna em numero e ela
    # chega aqui como '3.12608E+43' -- cruzaria com nada, ou pior, com a nota
    # errada. Chave torta para a rodada, em vez de virar buraco silencioso.
    tortas = [r[CHAVE] for r in registros
              if len(str(r[CHAVE]).strip()) != 44 or not str(r[CHAVE]).strip().isdigit()]
    if tortas:
        raise RuntimeError(
            f"{base.name}: {len(tortas)} chave(s) não têm 44 dígitos — "
            f"exemplo: {tortas[0]!r}.\n"
            "A chave NF-e não cabe em número: se o arquivo foi aberto e salvo no "
            "Excel, a coluna virou notação científica e o cruzamento erraria. "
            "Exporte de novo sem abrir/salvar, ou formate a coluna como Texto."
        )

    por_chave: dict[str, dict] = {}
    conflitos: list[str] = []
    for r in registros:
        chave = str(r[CHAVE]).strip()
        atual = por_chave.get(chave)
        if atual is None:
            por_chave[chave] = r
            continue
        a, b = atual[STATUS], r[STATUS]
        if a == b:
            continue
        if a == SEM_MANIFESTACAO:
            por_chave[chave] = r
        elif b != SEM_MANIFESTACAO:
            conflitos.append(f"{chave}: {a!r} ({atual[FILIAL].strip()}) x "
                             f"{b!r} ({r[FILIAL].strip()})")
    if conflitos:
        raise RuntimeError(
            f"{base.name}: {len(conflitos)} nota(s) com DOIS status de "
            "manifestação diferentes, e nenhum deles é 'SemManifestacao':\n  "
            + "\n  ".join(conflitos[:10])
            + "\nAté 01/09/2026 isso nunca aconteceu (todo conflito era um status "
              "real contra 'SemManifestacao'). Olhe a nota antes de escolher: "
              "qualquer regra automática aqui esconderia a divergência."
        )

    datas = [d for d in (_data(r[EMISSAO]) for r in registros) if d]
    resumo = {
        "arquivo": str(base),
        "linhas": len(registros),
        "notas": len(por_chave),
        "repetidas": len(registros) - len(por_chave),
        "de": min(datas) if datas else None,
        "ate": max(datas) if datas else None,
        "por_status": {s: sum(1 for r in por_chave.values() if r[STATUS] == s)
                       for s in sorted({r[STATUS] for r in por_chave.values()})},
    }
    return por_chave, resumo


def _valor(registro: dict, emissao, de, ate, e_servico: bool) -> str:
    """O que a coluna mostra para UMA nota. Quatro respostas possíveis."""
    if registro is not None:
        return registro[STATUS]
    # ⚠ NFS-e nao tem manifestacao do destinatario -- nao e omissao do arquivo, e
    # o instituto que nao existe para nota de servico. Medido: 0 das 366 NFS-e
    # do painel estao na base, e e o esperado.
    if e_servico:
        return NAO_SE_APLICA
    # ⚠ ESTE E O CASO QUE FAZ A COLUNA NAO MENTIR. A SEFAZ.xlsx e ACUMULATIVA e
    # a manifestacao vem por MES. Hoje (01/09/2026) as duas cobrem agosto e a
    # diferenca quase nao aparece -- 2 notas. Em outubro a SEFAZ tera ago+set+out
    # e o arquivo so outubro: sem esta distincao, o painel diria "sem
    # manifestacao" para centenas de notas de agosto que JA foram manifestadas.
    # O periodo sai do PROPRIO arquivo, e nao de um mes escrito aqui.
    if emissao is None or de is None or not (de <= emissao <= ate):
        return FORA_DO_PERIODO
    # Dentro do periodo e mesmo assim ausente: isso e anomalia de verdade.
    return NAO_ENCONTRADA


def anotar(linhas: list[dict], base: Path | None = None) -> dict:
    """Preenche `sefaz.MANIFESTACAO` em cada linha e devolve o resumo.

    Roda DEPOIS do `sefaz.ler_com_historico()` e ANTES do `sefaz.carregar()`:
    as tres abas que leem a SEFAZ enxergam a coluna, e o `sefaz_historico.json`
    continua guardando a linha pura. Manifestacao e fato de HOJE, relido a cada
    rodada -- congelar no historico faria a nota removida carregar para sempre o
    status que ela tinha no dia em que sumiu.
    """
    por_chave, resumo = ler(base)
    de, ate = resumo["de"], resumo["ate"]
    contagem: dict[str, int] = {}
    notas_vistas: set[str] = set()
    for linha in linhas:
        chave = linha[sefaz.CHAVE_NFE]
        registro = por_chave.get(chave)
        valor = _valor(registro, linha.get(sefaz.EMISSAO), de, ate,
                       linha[sefaz.ORIGEM] == sefaz.NFSE)
        linha[sefaz.MANIFESTACAO] = valor
        # O alerta e da NOTA e vai pela propria linha, como o `_removida_em`:
        # assim o `alertas_por_nota` o encontra sem que ninguem precise passar
        # mais um parametro pelos tres caminhos que chamam aquela funcao.
        if registro is not None and registro[STATUS] in GRAVES:
            justificativa = (registro.get(JUSTIFICATIVA) or "").strip()
            linha[sefaz.MANIFESTO_GRAVE] = (
                f"{registro[STATUS]}"
                + (f" — {justificativa}" if justificativa else "")
                + f" (filial {registro[FILIAL].strip()})")
        if chave not in notas_vistas:
            notas_vistas.add(chave)
            contagem[valor] = contagem.get(valor, 0) + 1
    resumo["anotadas"] = {k: contagem[k] for k in sorted(contagem)}
    resumo["notas_no_painel"] = len(notas_vistas)
    resumo["graves"] = sum(n for s, n in contagem.items() if s in GRAVES)
    return resumo


def marcar_sem_base(linhas: list[dict]) -> None:
    """Sem a base, a coluna diz que NAO FOI LIDA -- e nao fica vazia.

    Celula vazia seria lida como "esta nota nao tem manifestacao", que e uma
    afirmacao que ninguem fez: o arquivo nem foi aberto.
    """
    for linha in linhas:
        linha[sefaz.MANIFESTACAO] = SEM_BASE
