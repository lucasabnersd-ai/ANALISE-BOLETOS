# -*- coding: utf-8 -*-
"""PREMISSA 2: as notas do arquivo de manifestacao que a SEFAZ.xlsx nao tem.

04/09/2026, pedido do usuario: "inclua essa base no cruzamento... use o nome do
emissor pra cruzar e verifique se tem na SF1. Chame essa base de PREMISSA 2 no
painel".

O QUE ESTE MODULO FAZ E O QUE NAO FAZ
-------------------------------------
Ao contrario do `manifestacao.py` -- que le a MESMA PREMISSA.xlsx so para
ENRIQUECER as notas da SEFAZ com a manifestacao --, este e uma FONTE DE LINHAS:
as notas que estao aqui e NAO estao na SEFAZ.xlsx entram no painel como notas,
com `Origem = "PREMISSA 2"`, e seguem o mesmo caminho das outras: cruzamento com
a SF1 (esta lancada?) e, para as que nao estao, a aba NF x Pedido de Compra com
os pedidos provaveis do fornecedor.

⚠ QUAL ARQUIVO, E EM QUE FORMA -- as duas coisas mudam sem aviso.
No mesmo 04/09/2026 em que este modulo nasceu apontando para o
`MESMA PREMISSA - ORGANIZADA.xlsx` (aba `NOTAS`, 1.380 linhas), ele foi
reexportado e VOLTOU A SER o CSV cru de uma coluna so (aba com nome GUID, 584
notas), enquanto o organizado com 1.466 notas passou a ser o
`MESMA PREMISSA - CONSOLIDADA.xlsx`, que nao existia de manha. A rodada parou
com AVISO -- que e o comportamento certo --, e a licao virou codigo em tres
lugares:
  - `CANDIDATOS`: o arquivo e procurado por uma LISTA de nomes, do mais completo
    para o menos, e o `.cmd` nao fixa mais caminho nenhum;
  - `_tabela()`: a aba e escolhida pelo CABECALHO (a que tem mais colunas
    conhecidas), nunca pelo nome -- mesma regra do `sefaz._achar_aba()`;
  - `SINONIMOS`: os dois vocabularios das mesmas 14 colunas convivem
    (`Chave de Acesso` x `Chave`, `Número NF` x `Numero`...).
Medido depois: o leitor devolve 1.466 / 584 / 1.380 notas nos tres arquivos, com
emissao e valor lidos em 100% das linhas de cada um.

MEDIDO NA PRIMEIRA VERSAO (arquivo de 01/08 a 31/08, 1.380 linhas):
  - 56 chaves repetidas (a mesma nota vista por duas filiais, como no
    manifestacao.py), todas com 44 digitos limpos;
  - 858 sao notas de TERCEIROS (a Bracell e outros, em que a S&D so aparece como
    transportadora) -- o DF-e do TOTVS puxa tudo que cita o CNPJ do grupo em
    qualquer papel. Elas NAO entram: o corte e o mesmo da SEFAZ, CNPJ do
    destinatario na LISTAGEM EMPRESAS BIOFLOR;
  - das 466 do grupo, 413 ja estavam na SEFAZ.xlsx (pela chave de 44) e ficaram
    de fora -- ja estao no painel, com mais dado do que este arquivo tem;
  - sobraram 53 notas novas, 20 delas SEM lancamento na SF1.

⚠ POR QUE A CHAVE DE 44 E A IDENTIDADE, E NAO O NUMERO. A mesma nota pode estar
nas duas bases com o numero escrito diferente ("12765" x "000012765"), e o
numero da NFS-e comeca do 1 em cada prestador. A chave nao: `_digitos(chave)`
casa a mesma nota nos dois lados sem falso positivo, e foi assim que as 467
repetidas foram achadas.

⚠ O QUE ESTE ARQUIVO NAO TEM, E A LINHA NASCE SEM: itens, CFOP, duplicatas
(vencimento/valor da parcela), nome fantasia, tipo de operacao e o texto
`Informações` de onde a aba de pedidos le o numero do PC citado. Sem o texto, o
pedido dessas notas so pode ser achado por fornecedor + valor -- e e por isso
que a coluna `Numero PC Prováveis` importa mais ainda para elas. O `Status da
Nota` (Autorizada/Cancelada) tambem nao existe aqui e fica VAZIO -- inventar
"Autorizada" seria afirmar o que ninguem conferiu.

⚠ A MANIFESTACAO VEM DO PROPRIO ARQUIVO (coluna `Status Manifestação`), porque
ele E o arquivo de manifestacao. E gravada direto na linha; o `manifestacao.anotar`
roda ANTES desta injecao e por isso nao a sobrescreve.
"""
from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import re
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

import sefaz

_AQUI = Path(__file__).resolve()
_ANALISES = _AQUI.parents[1]

# ⚠ O ARQUIVO TEM DOIS NOMES E DUAS FORMAS, e os dois mudam sem aviso.
# 04/09/2026: o ORGANIZADA.xlsx foi reexportado e VOLTOU A SER o CSV cru (uma
# aba so, nome GUID, 585 linhas), enquanto quem passou a ter a aba NOTAS -- e as
# 1.466 notas -- foi o CONSOLIDADA.xlsx, que nasceu no mesmo dia. A leitura
# parou com AVISO, que e o certo; o que nao pode e depender do nome.
# A ordem abaixo e "o mais completo primeiro". `--base-premissa2` vence tudo.
CANDIDATOS = ("MESMA PREMISSA - CONSOLIDADA.xlsx", "MESMA PREMISSA - ORGANIZADA.xlsx")


def _primeira_que_existe() -> Path:
    for nome in CANDIDATOS:
        caminho = _ANALISES / nome
        if caminho.is_file():
            return caminho
    return _ANALISES / CANDIDATOS[0]


BASE_PADRAO = _primeira_que_existe()

# O que a coluna `Origem` mostra para estas linhas. ⚠ E um TERCEIRO valor ao lado
# de sefaz.NFE / sefaz.NFSE. Quem separa as linhas por origem (`sefaz.carregar`
# conta nfe/nfs, `manifestacao` pergunta "e servico?") continua funcionando: uma
# linha PREMISSA 2 nao e nem uma coisa nem outra, e nao entra nessas contas.
ORIGEM = "PREMISSA 2"

# ⚠ DOIS VOCABULARIOS PARA AS MESMAS 14 COLUNAS. O export cru chama
# `NomeFilial/Numero/Chave/CnpjCpfEmitente`; a versao organizada chama
# `Filial/Número NF/Chave de Acesso/CNPJ/CPF Emitente`. Aqui os dois apontam
# para o mesmo destino -- e por isso a leitura nao se importa com qual veio.
SINONIMOS = {
    "_filial_nome":       ("Filial", "NomeFilial"),
    sefaz.NUM_NF:         ("Número NF", "Numero"),
    "_serie":             ("Série", "Serie"),
    sefaz.EMISSAO:        ("Data Emissão", "DataEmissao"),
    sefaz.CHAVE_NFE:      ("Chave de Acesso", "Chave"),
    sefaz.NOME_EMIT:      ("Emissor", "Emissor"),
    sefaz.CNPJ_EMIT:      ("CNPJ/CPF Emitente", "CnpjCpfEmitente"),
    sefaz.NOME_DEST:      ("Destinatário", "Destinatario"),
    sefaz.CNPJ_DEST:      ("CNPJ/CPF Destinatário", "CnpjCpfDestinatario"),
    sefaz.VLR_TOTAL:      ("Valor NF (R$)", "ValorNF"),
    sefaz.MANIFESTACAO:   ("Status Manifestação", "StatusManifestacao"),
    sefaz.NAT_OP:         ("Natureza da Operação", "NaturezaOperacao"),
}
# Sem estas nao ha nota; o resto pode faltar que a linha nasce vazia.
OBRIGATORIOS = (sefaz.CHAVE_NFE, sefaz.NOME_EMIT, sefaz.CNPJ_EMIT,
                sefaz.CNPJ_DEST, sefaz.NUM_NF, sefaz.EMISSAO, sefaz.VLR_TOTAL)


def _digitos(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _data(valor):
    """A emissao, venha ela como for.

    Sao TRES formatos no mesmo campo: `datetime` de verdade (a aba NOTAS traz a
    coluna formatada), "2026-08-04 11:06:38" (a mesma aba quando o Excel a
    guardou como texto) e "04-08-2026T11:03:48AM-03:00" (o CSV cru). ⚠ O
    `sefaz._data` sozinho nao cobre os dois ultimos: ele tenta a string INTEIRA,
    e a hora colada faz todo strptime falhar -- a emissao viraria None e a nota
    sairia sem data, sem erro nenhum aparecer.
    """
    if isinstance(valor, (dt.datetime, dt.date)):
        return sefaz._data(valor)
    bruto = str(valor or "").strip()
    if not bruto:
        return None
    achado = sefaz._data(bruto)
    if achado:
        return achado
    cabeca = bruto[:10]                      # so a data; o resto e hora/fuso
    for formato in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(cabeca, formato).date()
        except ValueError:
            pass
    return None


def _abrir(caminho: Path):
    """Abre a planilha; se o Excel/OneDrive estiver segurando, usa uma copia.

    O mesmo remedio de sefaz.py / manifestacao.py: este arquivo vive aberto na
    tela de quem acabou de organiza-lo.
    """
    try:
        return load_workbook(caminho, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        pasta = tempfile.mkdtemp(prefix="premissa2_")
        shutil.copy2(caminho, Path(pasta) / caminho.name)
        return load_workbook(Path(pasta) / caminho.name, read_only=True,
                             data_only=True), pasta


def _tabela(wb) -> list[list[str]]:
    """As linhas do arquivo, venha ele como planilha ou como CSV numa coluna.

    ⚠ A ABA NAO E ESCOLHIDA PELO NOME -- e a que tem MAIS COLUNAS conhecidas no
    cabecalho. Foi o nome que quebrou a rodada de 04/09 ('NOTAS' virou um GUID),
    e a mesma licao ja esta no `sefaz._achar_aba`: o nome nao e identidade. A
    aba RESUMO, que tem 1 coluna de titulo, perde para a de dados por construcao.

    ⚠ O CSV E LIDO COM `csv.reader`, NUNCA COM `split(',')` -- mesma razao do
    manifestacao.py: ha virgula DENTRO de campo entre aspas, e o split parte a
    linha em pedacos a mais, jogando todos os campos seguintes uma casa para o
    lado. A chave passaria a ser o nome do emissor e o cruzamento erraria calado.
    """
    conhecidos = {n for nomes in SINONIMOS.values() for n in nomes}
    melhor, melhor_nota = [], -1
    for ws in wb.worksheets:
        if ws.max_column == 1:
            cruas = [c[0] for c in ws.iter_rows(values_only=True)]
            texto = "\n".join(str(c) for c in cruas
                              if c is not None and str(c).strip())
            tabela = [l for l in csv.reader(io.StringIO(texto)) if l]
        else:
            tabela = [["" if v is None else str(v).strip() for v in bruta]
                      for bruta in ws.iter_rows(values_only=True)
                      if not all(v in (None, "") for v in bruta)]
        if not tabela:
            continue
        nota = sum(1 for c in tabela[0] if str(c).strip() in conhecidos)
        if nota > melhor_nota:
            melhor, melhor_nota = tabela, nota
    return melhor


def ler(base: Path | None = None) -> list[dict]:
    """Todas as notas do arquivo, ja no vocabulario do painel. Sem filtro."""
    base = base or BASE_PADRAO
    if not base.exists():
        raise RuntimeError(
            f"A base PREMISSA 2 nao foi encontrada:\n  {base}\n"
            f"Procurei, nesta ordem: {' / '.join(CANDIDATOS)}. Confira se o "
            "OneDrive sincronizou, ou aponte com --base-premissa2.")
    wb, temporaria = _abrir(base)
    try:
        tabela = _tabela(wb)
    finally:
        wb.close()
        if temporaria:
            shutil.rmtree(temporaria, ignore_errors=True)
    if not tabela:
        raise RuntimeError(f"{base.name} esta vazio -- nada para cruzar.")

    cabecalho = [str(c).strip() if c is not None else "" for c in tabela[0]]
    # {indice da coluna: destino} -- o primeiro sinonimo que aparecer vence
    posicoes = {}
    for destino, nomes in SINONIMOS.items():
        for nome in nomes:
            if nome in cabecalho:
                posicoes[cabecalho.index(nome)] = destino
                break
    faltando = [d for d in OBRIGATORIOS if d not in posicoes.values()]
    if faltando:
        raise RuntimeError(
            f"{base.name} mudou de layout: nao achei coluna para "
            + " / ".join(repr(f) for f in faltando) + "\n  o arquivo tem: "
            + " / ".join(repr(c) for c in cabecalho if c)
            + "\nAcrescente o nome novo em SINONIMOS no premissa2.py.")

    linhas = []
    for bruta in tabela[1:]:
        linha = {sefaz.ORIGEM: ORIGEM}
        for i, destino in posicoes.items():
            valor = bruta[i] if i < len(bruta) else None
            if destino == sefaz.EMISSAO:
                linha[destino] = _data(valor)
            elif destino == sefaz.VLR_TOTAL:
                linha[destino] = sefaz._numero(valor)
            else:
                linha[destino] = sefaz._texto(valor)
        chave = _digitos(linha.get(sefaz.CHAVE_NFE))
        if len(chave) != 44:
            # rodape, celula vazia, ou chave que virou notacao cientifica ao
            # passar pelo Excel -- nada disso e nota
            continue
        linha[sefaz.CHAVE_NFE] = chave
        # ⚠ A filial vem como "003001 - BRASMAC..." -- o codigo e util para
        # conferir e fica num campo interno ("_", nao vai para a tela).
        linha["_filial_totvs"] = (linha.get("_filial_nome") or "").strip()[:6]
        # o que o arquivo NAO tem, e a linha nasce declarando que nao tem
        for vazio in (sefaz.STATUS_NOTA, sefaz.TIPO_OP, sefaz.FANTASIA,
                      sefaz.CFOP, sefaz.INFO):
            linha.setdefault(vazio, "")
        for nenhum in (sefaz.SAIDA, sefaz.VENC_DUP, sefaz.VLR_DUP, sefaz.VLR_ITEM):
            linha.setdefault(nenhum, None)
        # ⚠ Uma linha por NOTA: e o `Chave` que segura check e tratativa, e
        # `sefaz.PREFIXO` e o mesmo das outras para o painel tratar igual.
        linha[sefaz.CHAVE] = sefaz.PREFIXO + hashlib.sha1(
            f"{ORIGEM}|{chave}".encode("utf-8")).hexdigest()[:16]
        linhas.append(linha)
    return linhas


def linhas_novas(linhas_sefaz: list[dict], permitidos: dict,
                 base: Path | None = None) -> tuple[list[dict], dict]:
    """As notas PREMISSA 2 que ENTRAM no painel: do grupo e ausentes da SEFAZ.

    `permitidos` e o {CNPJ: razao} da LISTAGEM EMPRESAS BIOFLOR (o mesmo corte
    da SEFAZ). `linhas_sefaz` sao as linhas que o painel ja tem -- e delas que
    saem as chaves a NAO repetir.

    ⚠ A CHAVE REPETIDA DENTRO DO PROPRIO ARQUIVO (a mesma nota vista por duas
    filiais) vira UMA linha: vale a primeira, e a manifestacao segue a regra do
    manifestacao.py -- qualquer status vence `SemManifestacao`.
    """
    todas = ler(base)
    ja_no_painel = {_digitos(l.get(sefaz.CHAVE_NFE)) for l in linhas_sefaz}
    do_grupo = [l for l in todas if _digitos(l.get(sefaz.CNPJ_DEST)) in permitidos]

    por_chave: dict[str, dict] = {}
    for l in do_grupo:
        atual = por_chave.get(l[sefaz.CHAVE_NFE])
        if atual is None:
            por_chave[l[sefaz.CHAVE_NFE]] = l
        elif (atual.get(sefaz.MANIFESTACAO) or "") == "SemManifestacao" \
                and (l.get(sefaz.MANIFESTACAO) or "") not in ("", "SemManifestacao"):
            por_chave[l[sefaz.CHAVE_NFE]] = l

    novas = [l for ch, l in por_chave.items() if ch not in ja_no_painel]
    resumo = {
        "arquivo": str(base or BASE_PADRAO),
        "nome": (base or BASE_PADRAO).name,
        "linhas": len(todas),
        "do_grupo": len(por_chave),
        "terceiros": len(todas) - len(do_grupo),
        "ja_na_sefaz": len(por_chave) - len(novas),
        "novas": len(novas),
    }
    return novas, resumo
