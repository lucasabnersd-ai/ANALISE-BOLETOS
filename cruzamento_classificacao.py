# -*- coding: utf-8 -*-
"""Aba "Cruzamento Classificacao".

Pega os titulos que o pessoal ACABOU DE CLASSIFICAR no TOTVS (base SF1) e
procura, para cada um, o boleto correspondente entre os boletos que ainda NAO
acharam titulo -- os do DDA (aba "Futuros NF Nao Associada") e os do e-mail do
Itau (planilha COPIAR E COLAR BOLETOS PENDENTES).

Quem entra da SF1
-----------------
    Status (coluna BU) == "A"   e   Dt.Digitacao (coluna X) == dia util anterior

⚠ "dia util anterior" aqui NAO e so `hoje - 1`:

1. Fim de semana: segunda-feira olha para sexta.
2. FERIADO nao tem calendario nesta maquina. Em vez de chutar, o modulo cai
   para o DIA MAIS RECENTE QUE TENHA TITULOS classificados. Num feriado o dia
   vem vazio e ele recua sozinho -- sem lista de feriados para manter.
3. Se o painel ficar dias sem rodar, pegar so o ultimo dia util PERDERIA os
   dias do meio. Por isso o historico guarda ate onde ja foi lido
   (`ultimo_dia`) e a carga seguinte traz TUDO o que foi classificado depois
   disso. Na primeira execucao, so o ultimo dia util.

Acumula, nao repete
-------------------
Igual a aba do Itau ([[pendentes_itau]]): cada titulo entra UMA vez, guarda
"Entrou em" e continua no painel nos dias seguintes. A chave e o "Campo UUID"
da SF1 (a NF no TOTVS), com prefixo para nao colidir com os UUIDs das outras
abas na tabela de marcacoes.

⚠ O CRUZAMENTO E REFEITO A CADA CARGA, de proposito: o titulo fica guardado,
mas o boleto dele pode aparecer amanha. Um titulo que hoje esta "SEM BOLETO"
passa a "MATCH FORTE" sozinho quando o boleto entrar na base -- sem isso a aba
congelaria a resposta do primeiro dia.

As regras de cruzamento
-----------------------
Mesmo espirito do `consolidar_bases_boletos.py`: cada criterio vale um BIT
distinto, entao a soma diz exatamente o que bateu. O que muda e o acervo (aqui
sao duas origens de boleto) e o fato de a NF poder ter VARIAS PARCELAS.

    NF                8   numero da nota, so digitos e sem zeros a esquerda
    CNPJ              4   raiz de 8 digitos -- e a identidade real do fornecedor
    Valor             2   valor do boleto x valor da NF (ou soma das parcelas)
    Vencimento        1   1o vencimento da NF x vencimento do boleto
    Nome fornecedor   1   so quando nao ha CNPJ dos dois lados para comparar

Nome de fornecedor entra como DESEMPATE, nunca sozinho: o Itau trunca o
beneficiario em 30 caracteres e escreve a mesma empresa de varios jeitos, entao
nome batendo prova pouco. CNPJ e NF e que decidem.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import shutil
import tempfile
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

# Mesmo criterio do sc7.py e do sefaz.py: este arquivo mora em
# ...\LUCAS ABNER ARAUJO\<AUTOMACOES LUCAS>\ANALISES BOLETOS\PAINEL ANALISE
# BOLETOS\, entao quatro niveis acima esta a raiz -- e isso vale nas DUAS
# maquinas, que montam a mesma biblioteca do OneDrive com nomes diferentes.
# ⚠ Ate 31/08/2026 aqui so havia o caminho fixo do usuario "lucas". Era o unico
# default sem derivacao entre as bases da aba NF x Pedido de Compra, e o mais
# caro: sem a SF1 o gerar_painel.py pula de uma vez a aba SEFAZ x SF1 E a de
# pedidos de compra.
BASE_PADRAO = Path(__file__).resolve().parents[3] / "BASES GENERICOS" / "SF1.xlsx"
if not BASE_PADRAO.exists():
    BASE_PADRAO = Path(
        r"C:\Users\lucas\OneDrive - Grupo S&D\Arquivos de Gabriella Karla Oliveira Milas - FINANCEIRO COMPARTILHADO"
        r"\LUCAS ABNER ARAUJO\BASES GENERICOS\SF1.xlsx"
    )

PREFIXO = "SF1:"

# Colunas da SF1 por POSICAO. A planilha tem 220 colunas com nomes repetidos
# ("Peso Liquido" aparece duas vezes) -- casar por nome seria pior. As letras
# sao as que o usuario informou: Status = BU, Dt.Digitacao = X.
SF1_FILIAL = 0        # A
SF1_NUMERO = 1        # B  numero da NF
SF1_SERIE = 2         # C
SF1_FORNECEDOR = 3    # D  codigo do fornecedor = raiz do CNPJ
SF1_LOJA = 4          # E
SF1_RAZAO = 5         # F  "Descricao" no cabecalho, mas e a razao social
SF1_EMISSAO = 8       # I
SF1_PRIMEIRO_VENC = 9   # J
SF1_TIPO_PGTO = 10    # K
SF1_VLR_BRUTO = 19    # T
SF1_DIGITACAO = 23    # X   <- condicao do usuario
SF1_ESPECIE = 51      # AZ
SF1_STATUS = 72       # BU  <- condicao do usuario
SF1_CHAVE_NFE = 101   # CX
SF1_UUID = 214        # HG

STATUS_CLASSIFICADO = "A"

# Nomes de saida. "Vencimento" tem de se chamar assim: a ordem padrao, o filtro
# de periodo e a formatacao de data do painel procuram essa chave.
CHAVE = "Chave"
SITUACAO = "Situação"
ENTROU = "Entrou em"
CLASSIFICADO = "Classificado em"
FILIAL = "Filial"
NF = "Nº NF"
SERIE = "Série"
FORNECEDOR = "Fornecedor"
RAZAO = "Razão Social"
VLR_TITULO = "Vlr.Título"
VENCIMENTO = "Vencimento"
EMISSAO = "Emissão"
UUID_NF = "Campo UUID"
# Tipo de pagamento gravado na NF (SF1, coluna K). O nome carrega o "SF1" de
# proposito: na aba fundida existe uma coluna VIRTUAL "@tipo_pgto" (o tipo que
# vem da SE2) e `chave_de` reduziria as duas a `tipo_pgto` -- uma apagaria a
# outra. Esta coluna nao e desenhada: o valor viaja para o painel pelo mapa
# `se2.tipos`, que e por onde o navegador ja preenche essa celula.
TIPO_PGTO = "Tipo Pgto SF1"
# lado do boleto
NF_BOLETO = "NF/Doc Boleto"
FORN_BOLETO = "Fornecedor Boleto"
VLR_BOLETO = "Valor Boleto"
VENC_BOLETO = "Vencimento Boleto"
FONTE = "Fonte Boleto"
PARCELA = "Parcela"
CRITERIO = "Critério"
CANDIDATOS = "Boletos"
ALERTA = "Alerta Fatura"
DIGITAVEL = "Linha Digitável"

CABECALHO = [CHAVE, SITUACAO, ENTROU, CLASSIFICADO, FILIAL, NF, NF_BOLETO,
             SERIE, FORNECEDOR, RAZAO, FORN_BOLETO, VLR_TITULO, VLR_BOLETO,
             VENCIMENTO, VENC_BOLETO, EMISSAO, PARCELA, CANDIDATOS, CRITERIO,
             FONTE, ALERTA, DIGITAVEL, UUID_NF, TIPO_PGTO]

# Selos. Curtos de proposito: a celula corta o que nao couber.
FORTE = "MATCH FORTE"
PROVAVEL = "MATCH PROVÁVEL"
REVISAR = "REVISAR"
AMBIGUO = "AMBÍGUO"
SEM_BOLETO = "SEM BOLETO"

# ⚠ Os pesos sao BITS DISTINTOS -- e a soma tem de ordenar do mais forte para o
# mais fraco, porque o melhor candidato e escolhido por `max(nota)`. Se "nome"
# valesse mais que "NF", um homonimo ganharia da nota certa.
PESO_NF = 16
PESO_CNPJ = 8
PESO_VALOR = 4
PESO_VENCIMENTO = 2
PESO_NOME = 1

NOME_CRITERIO = ((PESO_NF, "NF"), (PESO_CNPJ, "CNPJ"), (PESO_VALOR, "valor"),
                 (PESO_VENCIMENTO, "vencimento"), (PESO_NOME, "nome"))

# Centavos de tolerancia no confronto de valor.
TOLERANCIA = 0.02
# Janela do vencimento. Exato vale ponto cheio; ate aqui ainda conta como
# "perto" para desempatar candidatos, sem virar confirmacao.
DIAS_PERTO = 5

# Datas absurdas existem na SF1: notas de remessa vem com "1o Venc" no ano
# 5023. Passar isso adiante estragaria a ordenacao do painel.
ANO_MAX = 2100


# ---------------------------------------------------------------- utilitarios

def _sem_acento(texto) -> str:
    normalizado = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (dt.datetime, dt.date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return re.sub(r"\s+", " ", str(valor).replace("\xa0", " ")).strip()


def _data(valor):
    if isinstance(valor, dt.datetime):
        valor = valor.date()
    if isinstance(valor, dt.date):
        return valor if valor.year <= ANO_MAX else None
    bruto = _texto(valor)
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            achada = dt.datetime.strptime(bruto, formato).date()
        except ValueError:
            continue
        return achada if achada.year <= ANO_MAX else None
    return None


def _moeda(valor):
    if valor in (None, "", "-"):
        return None
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    txt = str(valor).replace("R$", "").replace("\xa0", "").replace(" ", "").strip()
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    txt = re.sub(r"[^0-9.\-]", "", txt)
    try:
        return round(float(txt), 2)
    except ValueError:
        return None


def _nf(valor) -> str:
    """Numero da nota comparavel: so digitos, sem zeros a esquerda.

    '000000031', '031' e 31 viram '31' -- e o mesmo tratamento que o
    consolidar_bases_boletos.py da ao documento.
    """
    digitos = re.sub(r"\D", "", _texto(valor)).lstrip("0")
    return digitos


def _raiz_cnpj(valor) -> str:
    """Os 8 primeiros digitos do CNPJ: identificam a empresa, nao a filial.

    Filiais diferentes do mesmo fornecedor emitem boleto pelo mesmo CNPJ raiz;
    exigir os 14 digitos perderia esses casos. Para CPF (11 digitos) o corte em
    8 nao serve, entao ele volta inteiro.
    """
    digitos = re.sub(r"\D", "", _texto(valor))
    if len(digitos) == 11:
        return digitos
    return digitos[:8] if len(digitos) >= 8 else ""


def _nome(valor) -> str:
    return re.sub(r"[^0-9A-Z]", "", _sem_acento(valor).upper())


def _nomes_batem(a: str, b: str) -> bool:
    """Comparacao por prefixo, porque o Itau trunca o beneficiario em 30."""
    if not a or not b:
        return False
    corte = min(len(a), len(b), 10)
    return corte >= 6 and a[:corte] == b[:corte]


def dia_util_anterior(hoje: dt.date) -> dt.date:
    """Dia util imediatamente anterior. Sabado e domingo nao contam.

    Feriado nao entra aqui (nao ha calendario) -- quem resolve isso e o recuo
    por dia vazio em `dias_a_processar`.
    """
    dia = hoje - dt.timedelta(days=1)
    while dia.weekday() >= 5:          # 5 = sabado, 6 = domingo
        dia -= dt.timedelta(days=1)
    return dia


# ------------------------------------------------------------------ leitura

def _abrir(caminho: Path):
    """Abre a planilha; se o Excel/OneDrive estiver segurando, usa uma copia."""
    try:
        return load_workbook(caminho, read_only=True, data_only=True), None
    except (PermissionError, OSError):
        pasta = tempfile.mkdtemp(prefix="cruzamento_")
        shutil.copy2(caminho, Path(pasta) / caminho.name)
        return load_workbook(Path(pasta) / caminho.name, read_only=True,
                             data_only=True), pasta


def ler_sf1(base: Path) -> list[dict]:
    """Todos os titulos com Status 'A' e data de digitacao valida.

    O filtro por DIA fica de fora de proposito: quem decide quais dias entram e
    o historico, e para isso ele precisa enxergar o que existe na base.
    """
    wb, temporaria = _abrir(base)
    try:
        ws = wb[wb.sheetnames[0]]
        brutas = ws.iter_rows(values_only=True)
        next(brutas, None)                       # cabecalho
        titulos = []
        for linha in brutas:
            if len(linha) <= SF1_UUID:
                continue
            if _texto(linha[SF1_STATUS]).upper() != STATUS_CLASSIFICADO:
                continue
            digitacao = _data(linha[SF1_DIGITACAO])
            if not digitacao:
                continue
            uuid = _texto(linha[SF1_UUID])
            titulos.append({
                "dia": digitacao,
                "uuid": uuid,
                FILIAL: _texto(linha[SF1_FILIAL]),
                NF: _texto(linha[SF1_NUMERO]),
                SERIE: _texto(linha[SF1_SERIE]),
                FORNECEDOR: _texto(linha[SF1_FORNECEDOR]),
                RAZAO: _texto(linha[SF1_RAZAO]),
                VLR_TITULO: _moeda(linha[SF1_VLR_BRUTO]),
                VENCIMENTO: _data(linha[SF1_PRIMEIRO_VENC]),
                EMISSAO: _data(linha[SF1_EMISSAO]),
                TIPO_PGTO: _texto(linha[SF1_TIPO_PGTO]),
                "especie": _texto(linha[SF1_ESPECIE]),
                "chave_nfe": _texto(linha[SF1_CHAVE_NFE]),
                "loja": _texto(linha[SF1_LOJA]),
            })
        return titulos
    finally:
        wb.close()
        if temporaria:
            shutil.rmtree(temporaria, ignore_errors=True)


def chave_do_titulo(titulo: dict) -> str:
    """Identidade do titulo no painel.

    O "Campo UUID" da SF1 e a chave natural (e o registro da NF no TOTVS). Ele
    vem vazio em algumas linhas antigas; nesses casos a identidade e montada com
    filial + NF + serie + fornecedor, que juntos nao se repetem.
    """
    if titulo["uuid"]:
        return PREFIXO + titulo["uuid"]
    partes = [titulo[FILIAL], _nf(titulo[NF]), _nome(titulo[SERIE]),
              _nome(titulo[FORNECEDOR]), titulo["loja"]]
    return PREFIXO + "-".join(partes)


# ------------------------------------------------------------------- boletos

def boletos_do_dda(linhas_futuros: list[dict]) -> list[dict]:
    """Boletos do DDA que nao acharam titulo (aba "Futuros NF Nao Associada")."""
    acervo = []
    for r in linhas_futuros:
        acervo.append({
            "fonte": _texto(r.get("Fonte Boleto")) or "DDA",
            "nf": _nf(r.get("NF Normalizada") or r.get("NF/Doc Original")),
            "nf_exibida": _texto(r.get("NF/Doc Original")),
            "cnpj": _raiz_cnpj(r.get("CNPJ/CPF")),
            "nome": _nome(r.get("Fornecedor")),
            "nome_exibido": _texto(r.get("Fornecedor")),
            "valor": _moeda(r.get("Valor (R$)")),
            "vencimento": _data(r.get("Vencimento")),
            "parcela": _texto(r.get("Parcela")),
            "total_parcelas": _texto(r.get("Total Parcelas")),
            "digitavel": _texto(r.get("Linha Digitável")),
            "alerta": "",
        })
    return acervo


def boletos_do_itau(historico_itau: dict) -> list[dict]:
    """Boletos pendentes vindos do e-mail do Itau (planilha copiar-e-colar).

    Os que ja sairam da planilha do banco (provavelmente pagos) ficam de fora:
    associar um titulo novo a um boleto que ja nao existe seria enganoso.
    """
    acervo = []
    for guardado in (historico_itau or {}).get("boletos", {}).values():
        if guardado.get("saiu_em"):
            continue
        linha = guardado.get("linha") or {}
        acervo.append({
            "fonte": "ITAU_EMAIL",
            "nf": _nf(linha.get("Nº Doc.")),
            "nf_exibida": _texto(linha.get("Nº Doc.")),
            "cnpj": _raiz_cnpj(linha.get("CPF/CNPJ")),
            "nome": _nome(linha.get("Beneficiário")),
            "nome_exibido": _texto(linha.get("Beneficiário")),
            "valor": _moeda(linha.get("Até Venc.")),
            "vencimento": _data(linha.get("Vencimento")),
            "parcela": "",
            "total_parcelas": "",
            "digitavel": "",
            "alerta": _texto(linha.get("Alerta Fatura")),
        })
    return acervo


def indexar(acervo: list[dict]) -> dict:
    """Indice por NF e por CNPJ -- so assim o cruzamento nao vira O(n x m).

    Sao ~800 boletos contra os titulos do dia; sem indice ainda rodaria, mas o
    acervo cresce e o indice tambem serve para achar candidato SEM NF (pelo
    CNPJ), que e o caso dos boletos do e-mail do Itau.
    """
    por_nf, por_cnpj = defaultdict(list), defaultdict(list)
    for boleto in acervo:
        if boleto["nf"]:
            por_nf[boleto["nf"]].append(boleto)
        if boleto["cnpj"]:
            por_cnpj[boleto["cnpj"]].append(boleto)
    return {"nf": dict(por_nf), "cnpj": dict(por_cnpj)}


# ---------------------------------------------------------------- cruzamento

def _pontuar(titulo: dict, boleto: dict, total_grupo=None) -> int:
    """Bits do que bateu entre o titulo e o boleto."""
    nota = 0
    nf_titulo = _nf(titulo[NF])
    if nf_titulo and boleto["nf"] == nf_titulo:
        nota += PESO_NF

    cnpj_titulo = _raiz_cnpj(titulo[FORNECEDOR])
    tem_cnpj = bool(cnpj_titulo and boleto["cnpj"])
    if tem_cnpj and cnpj_titulo == boleto["cnpj"]:
        nota += PESO_CNPJ
    elif not tem_cnpj and _nomes_batem(_nome(titulo[RAZAO]), boleto["nome"]):
        # nome so entra quando nao ha CNPJ dos dois lados: ele prova pouco
        nota += PESO_NOME

    valor = titulo[VLR_TITULO]
    if valor is not None and boleto["valor"] is not None:
        # parcela unica bate com o total da NF; havendo parcelas, o que tem de
        # fechar e a SOMA do grupo -- e por isso que `total_grupo` existe
        alvo = total_grupo if total_grupo is not None else boleto["valor"]
        if abs(alvo - valor) <= TOLERANCIA:
            nota += PESO_VALOR

    if titulo[VENCIMENTO] and boleto["vencimento"] == titulo[VENCIMENTO]:
        nota += PESO_VENCIMENTO
    return nota


def _distancia_venc(titulo: dict, boleto: dict) -> int:
    if not titulo[VENCIMENTO] or not boleto["vencimento"]:
        return 10**6
    return abs((boleto["vencimento"] - titulo[VENCIMENTO]).days)


def _criterio(nota: int) -> str:
    partes = [nome for peso, nome in NOME_CRITERIO if nota & peso]
    return " + ".join(partes)


def aceitavel(nota: int) -> bool:
    """O minimo para o boleto ser considerado candidato do titulo.

    ⚠ Sem esta barra o cruzamento vira ruido: o mesmo fornecedor tem varios
    boletos em aberto, entao "so o CNPJ bateu" acontece o tempo todo e NAO e
    associacao nenhuma -- era o que enchia a aba de AMBIGUO na primeira versao.
    Exige-se a NF, ou o fornecedor com pelo menos uma confirmacao de valor/data.
    """
    tem_nf = bool(nota & PESO_NF)
    tem_forn = bool(nota & (PESO_CNPJ | PESO_NOME))
    confirma = bool(nota & (PESO_VALOR | PESO_VENCIMENTO))
    return tem_nf or (tem_forn and confirma)


def _selo(nota: int, empatados: int) -> str:
    """Traduz a nota em selo. Pressupoe que a nota ja passou por `aceitavel`.

    A regra segue o `escolher_linha_digitavel`: numero exato sozinho nao basta,
    precisa de confirmacao. Aqui o CNPJ e a confirmacao mais forte.
    """
    tem_nf = bool(nota & PESO_NF)
    tem_cnpj = bool(nota & PESO_CNPJ)
    confirmacoes = sum(bool(nota & peso) for peso in
                       (PESO_CNPJ, PESO_VALOR, PESO_VENCIMENTO))
    if empatados > 1:
        return AMBIGUO
    if tem_nf and tem_cnpj and confirmacoes >= 2:
        return FORTE
    if tem_nf and confirmacoes >= 1:
        return PROVAVEL
    if tem_cnpj and (nota & PESO_VALOR) and (nota & PESO_VENCIMENTO):
        # sem NF, mas fornecedor + valor + data batendo e forte o bastante
        return PROVAVEL
    return REVISAR


def cruzar(titulo: dict, indice: dict) -> dict:
    """Melhor boleto para este titulo, com o selo e o criterio do que bateu."""
    nf_titulo = _nf(titulo[NF])
    cnpj_titulo = _raiz_cnpj(titulo[FORNECEDOR])

    candidatos = list(indice["nf"].get(nf_titulo, [])) if nf_titulo else []
    vistos = {id(b) for b in candidatos}
    for boleto in indice["cnpj"].get(cnpj_titulo, []) if cnpj_titulo else []:
        if id(boleto) not in vistos:
            candidatos.append(boleto)
            vistos.add(id(boleto))

    if not candidatos:
        return {"selo": SEM_BOLETO, "boleto": None, "criterio": "",
                "candidatos": 0, "nota": 0}

    # Parcelas da MESMA nota somam: o boleto vale 1/3 do titulo e, isolado, o
    # valor nunca bateria. Agrupar por NF+CNPJ e o que permite conferir o total.
    grupos = defaultdict(list)
    for boleto in candidatos:
        grupos[(boleto["nf"], boleto["cnpj"])].append(boleto)
    total_do_grupo = {}
    for chave, grupo in grupos.items():
        valores = [b["valor"] for b in grupo if b["valor"] is not None]
        total_do_grupo[chave] = round(sum(valores), 2) if len(grupo) > 1 and valores else None

    avaliados = []
    for boleto in candidatos:
        grupo = (boleto["nf"], boleto["cnpj"])
        nota = _pontuar(titulo, boleto, total_do_grupo.get(grupo))
        if aceitavel(nota):
            avaliados.append((nota, boleto))

    if not avaliados:
        return {"selo": SEM_BOLETO, "boleto": None, "criterio": "",
                "candidatos": 0, "nota": 0}
    melhor_nota = max(n for n, _ in avaliados)

    empatados = [b for n, b in avaliados if n == melhor_nota]
    # Parcelas nao sao empate de verdade: sao a mesma nota, fatiada. Desempata
    # pela parcela 1 (ou pelo vencimento mais proximo do 1o venc do titulo).
    parcelas = {b["parcela"] for b in empatados}
    mesmo_documento = len({(b["nf"], b["cnpj"]) for b in empatados}) == 1
    ambiguidade = len(empatados) if not (mesmo_documento and len(parcelas) > 1) else 1

    escolhido = min(empatados, key=lambda b: (_distancia_venc(titulo, b),
                                              _texto(b["parcela"])))
    return {
        "selo": _selo(melhor_nota, ambiguidade),
        "boleto": escolhido,
        "criterio": _criterio(melhor_nota),
        "candidatos": len(empatados),
        "nota": melhor_nota,
    }


# ------------------------------------------------------------------ historico

def _ler_historico(arquivo: Path) -> dict:
    if not arquivo.exists():
        return {"versao": 1, "ultimo_dia": None, "titulos": {}}
    try:
        guardado = json.loads(arquivo.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        # o mesmo cuidado do pendentes_itau: historico ilegivel nao pode virar
        # historico vazio em silencio -- seria o painel esquecendo tudo
        raise RuntimeError(
            f"{arquivo.name} esta ilegivel. Nada foi gerado -- restaure o arquivo "
            "(ou apague-o de proposito, sabendo que todo titulo voltara como novo)."
        )
    guardado.setdefault("titulos", {})
    guardado.setdefault("ultimo_dia", None)
    return guardado


def dias_a_processar(titulos: list[dict], historico: dict,
                     hoje: dt.date) -> tuple[list[dt.date], dt.date | None]:
    """Quais dias de digitacao entram nesta carga.

    - Primeira vez: so o ultimo dia util. Se ele estiver vazio (feriado, ou o
      TOTVS ainda nao exportou), recua para o dia mais recente COM titulos.
    - Depois: tudo o que foi classificado apos o ultimo dia ja lido, ate o
      limite do dia util anterior. E o que evita perder os dias do meio quando
      o painel fica sem rodar.
    """
    disponiveis = sorted({t["dia"] for t in titulos})
    if not disponiveis:
        return [], None

    limite = dia_util_anterior(hoje)
    ate = [d for d in disponiveis if d <= limite]
    if not ate:
        return [], limite

    ultimo = historico.get("ultimo_dia")
    if not ultimo:
        # so o dia mais recente que realmente tem titulos (feriado cai aqui)
        return [ate[-1]], limite

    corte = dt.date.fromisoformat(ultimo)
    return [d for d in ate if d > corte], limite


def atualizar_historico(base: Path, arquivo: Path, linhas_futuros: list[dict],
                        historico_itau: dict,
                        hoje: dt.date | None = None) -> tuple[dict, dict]:
    """Le a SF1, acrescenta os titulos novos e REFAZ o cruzamento de todos."""
    hoje = hoje or dt.date.today()
    historico = _ler_historico(arquivo)
    guardados = historico["titulos"]

    da_base = ler_sf1(base)
    dias, limite = dias_a_processar(da_base, historico, hoje)
    do_dia = [t for t in da_base if t["dia"] in set(dias)]

    novos = 0
    for titulo in do_dia:
        chave = chave_do_titulo(titulo)
        gravavel = {k: (v.isoformat() if isinstance(v, dt.date) else v)
                    for k, v in titulo.items() if k != "dia"}
        gravavel["dia"] = titulo["dia"].isoformat()
        if chave in guardados:
            # ja conhecido: atualiza os dados da NF, preserva a data de entrada
            guardados[chave]["titulo"] = gravavel
            continue
        guardados[chave] = {"entrou_em": hoje.isoformat(), "titulo": gravavel}
        novos += 1

    # ⚠ Preenchimento RETROATIVO do Tipo Pgto. A NF so e regravada no dia em que
    # e lida (o `continue` logo acima), entao uma coluna NOVA nasceria vazia em
    # todo mundo que ja estava no historico -- ela so passaria a valer para quem
    # chegasse depois, e a aba mostraria 95 celulas em branco sem explicacao.
    # A base ja foi lida INTEIRA em `ler_sf1` (sem filtro de dia), entao a volta
    # custa nada e o problema se resolve sozinho na primeira rodada.
    da_base_por_chave = {chave_do_titulo(t): t for t in da_base}
    for chave, guardado in guardados.items():
        if guardado["titulo"].get(TIPO_PGTO):
            continue
        origem = da_base_por_chave.get(chave)
        if origem and origem.get(TIPO_PGTO):
            guardado["titulo"][TIPO_PGTO] = origem[TIPO_PGTO]

    # O cruzamento e refeito para TODOS, nao so para os novos: o boleto de um
    # titulo antigo pode ter chegado hoje.
    indice = indexar(boletos_do_dda(linhas_futuros) + boletos_do_itau(historico_itau))
    contagem = defaultdict(int)
    for guardado in guardados.values():
        titulo = _reconstruir(guardado["titulo"])
        achado = cruzar(titulo, indice)
        guardado["match"] = _gravavel_match(achado)
        contagem[achado["selo"]] += 1

    if dias:
        historico["ultimo_dia"] = max(dias).isoformat()
    elif not historico.get("ultimo_dia") and limite:
        historico["ultimo_dia"] = limite.isoformat()
    historico["versao"] = 1
    historico["atualizado_em"] = dt.datetime.now().isoformat(timespec="seconds")
    historico["base"] = str(base)

    resumo = {
        "dias_lidos": [d.isoformat() for d in dias],
        "limite": limite.isoformat() if limite else "",
        "novos": novos,
        "total": len(guardados),
        "com_boleto": sum(v for k, v in contagem.items() if k not in ("", SEM_BOLETO)),
        "sem_boleto": contagem.get(SEM_BOLETO, 0) + contagem.get("", 0),
        "selos": dict(contagem),
    }
    return historico, resumo


def _reconstruir(gravavel: dict) -> dict:
    """JSON -> dict do titulo, com as datas de volta como `date`."""
    titulo = dict(gravavel)
    for campo in (VENCIMENTO, EMISSAO):
        titulo[campo] = _data(titulo.get(campo))
    titulo["dia"] = _data(titulo.get("dia"))
    return titulo


def _gravavel_match(achado: dict) -> dict:
    boleto = achado["boleto"]
    return {
        "selo": achado["selo"],
        "criterio": achado["criterio"],
        "candidatos": achado["candidatos"],
        "boleto": None if not boleto else {
            "nf": boleto["nf_exibida"],
            "nome": boleto["nome_exibido"],
            "valor": boleto["valor"],
            "vencimento": boleto["vencimento"].isoformat() if boleto["vencimento"] else "",
            "parcela": boleto["parcela"],
            "total_parcelas": boleto["total_parcelas"],
            "fonte": boleto["fonte"],
            "digitavel": boleto["digitavel"],
            "alerta": boleto["alerta"],
        },
    }


def gravar_historico(historico: dict, arquivo: Path) -> None:
    """Grava num temporario e so entao substitui: uma queda no meio da escrita
    nao pode deixar o historico pela metade."""
    arquivo.parent.mkdir(parents=True, exist_ok=True)
    provisorio = arquivo.with_suffix(arquivo.suffix + ".novo")
    provisorio.write_text(json.dumps(historico, ensure_ascii=False, indent=1),
                          encoding="utf-8")
    provisorio.replace(arquivo)


def montar_linhas(historico: dict) -> list[tuple]:
    """Historico -> tuplas na ordem de CABECALHO, prontas para o `montar_aba`."""
    registros = []
    for chave, guardado in historico["titulos"].items():
        t = guardado["titulo"]
        m = guardado.get("match") or {}
        b = m.get("boleto") or {}
        parcela = b.get("parcela") or ""
        total = b.get("total_parcelas") or ""
        registros.append({
            CHAVE: chave,
            SITUACAO: m.get("selo") or SEM_BOLETO,
            ENTROU: _data(guardado.get("entrou_em")),
            CLASSIFICADO: _data(t.get("dia")),
            FILIAL: t.get(FILIAL, ""),
            NF: t.get(NF, ""),
            NF_BOLETO: b.get("nf", ""),
            SERIE: t.get(SERIE, ""),
            FORNECEDOR: t.get(FORNECEDOR, ""),
            RAZAO: t.get(RAZAO, ""),
            FORN_BOLETO: b.get("nome", ""),
            VLR_TITULO: t.get(VLR_TITULO),
            VLR_BOLETO: b.get("valor"),
            VENCIMENTO: _data(t.get(VENCIMENTO)),
            VENC_BOLETO: _data(b.get("vencimento")),
            EMISSAO: _data(t.get(EMISSAO)),
            PARCELA: f"{parcela}/{total}" if parcela and total else parcela,
            CANDIDATOS: m.get("candidatos") or "",
            CRITERIO: m.get("criterio", ""),
            FONTE: b.get("fonte", ""),
            ALERTA: b.get("alerta", ""),
            DIGITAVEL: b.get("digitavel", ""),
            UUID_NF: (t.get("uuid") or ""),
            # `.get` com padrao: o historico gravado antes desta coluna existir
            # nao tem a chave, e um KeyError aqui derrubaria a aba inteira.
            TIPO_PGTO: t.get(TIPO_PGTO, ""),
        })

    # sem boleto primeiro (e o que precisa de acao), depois por classificacao
    # mais recente e vencimento mais proximo
    ordem_selo = {SEM_BOLETO: 0, AMBIGUO: 1, REVISAR: 2, PROVAVEL: 3, FORTE: 4}
    registros.sort(key=lambda r: (ordem_selo.get(r[SITUACAO], 0),
                                  -(r[CLASSIFICADO] or dt.date.min).toordinal(),
                                  r[VENCIMENTO] or dt.date.max))
    return [tuple(r[coluna] for coluna in CABECALHO) for r in registros]


def carregar(base: Path, arquivo_historico: Path, linhas_futuros: list[dict],
             historico_itau: dict,
             hoje: dt.date | None = None) -> tuple[list[str], list[tuple], dict]:
    """Ponto de entrada: le a SF1, cruza com os boletos e devolve
    (cabecalho, linhas, resumo) para o gerador."""
    historico, resumo = atualizar_historico(
        base, arquivo_historico, linhas_futuros, historico_itau, hoje)
    gravar_historico(historico, arquivo_historico)
    return list(CABECALHO), montar_linhas(historico), resumo
